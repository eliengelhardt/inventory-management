import os
import pickle
import traceback
from math import ceil
import time
import requests
import urllib
from PIL import Image
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import string
import math
from random import *
import glob
import json
import copy
import threading
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Color, PatternFill, Font, Border, Protection
from openpyxl.workbook.views import BookView

excel_window_width = 28000
excel_window_height = 23000
path_to_excel_dir = "/Users/davidhammons/Documents/item_cost_outside_onedrive/item_costs_excel_files"


def labor_for_sku_labeling_calculator(item_weight_lbs, item_volume_inch):
    dimensional_weight = item_volume_inch / 139
    weight_to_use = item_weight_lbs
    if dimensional_weight > weight_to_use:
        weight_to_use = dimensional_weight
    time_per_item_in_hours = 0.002666 + (weight_to_use * 0.000916)
    cost_per_item = time_per_item_in_hours * 30
    return cost_per_item


def get_skus_and_quantity():
    name_of_file = "skus_to_estimate.xlsx"

    workbook = openpyxl.Workbook()
    view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
    workbook.views = view
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    sheet["A1"] = "SKU"
    sheet["B1"] = "Quantity"
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    workbook.save(name_of_file)

    while True:
        print("enter 'done' when adding skus and quantitys to skus_to_estimate.xlsx")
        x = input()
        if x == "done":
            break

    dict_to_return = {}
    workbook = openpyxl.load_workbook(name_of_file)
    sheet = workbook.active
    cur_row = 2
    while True:
        sku = sheet["A" + str(cur_row)].value
        if sku == None:
            break
        q = float(sheet["B" + str(cur_row)].value)
        dict_to_return[sku] = q
        cur_row += 1

    return dict_to_return


def get_total(dict_of_sku_to_quantity):
    total = 0

    pars_needed = {}
    for par_var in dict_of_sku_to_quantity:
        index_of_under = par_var.index("-")
        if index_of_under == -1:
            raise Exception('index_of_under == -1')
        par_num = int(par_var[0:index_of_under])
        pars_needed[par_num] = True

    dict_of_par_var_to_quantity = dict_of_sku_to_quantity
    # for sku in dict_of_sku_to_quantity:
    #     index_of_under = sku.rfind("-")
    #     if index_of_under == -1:
    #         raise Exception('index_of_under == -1')
    #     par_var_str = sku[0:index_of_under]
    #     dict_of_par_var_to_quantity[par_var_str] = dict_of_sku_to_quantity[sku]

    list_of_excel_files = []
    for file in glob.glob(path_to_excel_dir + "/*.xlsx"):
        list_of_excel_files.append(file)

    for file in list_of_excel_files:
        filename = Path(file).name
        index_of_under = filename.index("_")
        if index_of_under == -1:
            raise Exception('index_of_under == -1')
        par_num = int(filename[0:index_of_under])
        if par_num in pars_needed:

            workbook = openpyxl.load_workbook(file)

            for sheet in workbook:
                var_num = int(sheet.title)
                par_var = str(par_num) + "-" + str(var_num)
                if par_var in dict_of_par_var_to_quantity:
                    quantity = dict_of_par_var_to_quantity[par_var]
                    weight_lbs = float(sheet["E2"].value)
                    length_in = float(sheet["E4"].value)
                    width_in = float(sheet["E5"].value)
                    height_in = float(sheet["E6"].value)
                    volume_inch = length_in * width_in * height_in
                    total = total + (quantity * labor_for_sku_labeling_calculator(weight_lbs, volume_inch))
    return total


if __name__ == '__main__':
    dict_of_sku_to_quantity = get_skus_and_quantity()

    tot = get_total(dict_of_sku_to_quantity)
    print("$" + str(tot))
