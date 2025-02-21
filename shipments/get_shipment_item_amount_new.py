import os
import pickle
import shutil
from distutils.dir_util import copy_tree
import requests
import time
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Color, PatternFill, Font, Border, Protection
from openpyxl.workbook.views import BookView
import math as m
import json
import urllib
import xml.etree.ElementTree as ET
from xml.dom import minidom
from PyPDF2 import PdfWriter, PdfReader
import io
from reportlab.pdfgen import canvas
import cv2
from pyzbar.pyzbar import decode
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics import renderPDF
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode.code128 import Code128
from pdf2image import convert_from_path
from reportlab.lib import colors
import pytesseract
import sys
import easyocr
import copy
import glob



excel_window_width = 28000
excel_window_height = 23000

full_path = os.getcwd()
parent = os.path.abspath(os.path.join(full_path, os.pardir))
all_pickle_path = parent + "/all_pickle_files"
all_pickle_path_copy = parent + "/all_pickle_files_copy"

loc_of_amazon_fba_sku_dict = all_pickle_path + "/amazon_fba_sku_info_for_shipments.pkl"
loc_of_supply_name_file = all_pickle_path + "/amazon_sku_to_parent_variation_num_dict.pkl"

loc_of_sku_inventory_file = all_pickle_path + "/sku_inventory_level.pkl"

loc_of_supply_inventory_file = all_pickle_path + "/supplies_inventory_level.pkl"
loc_of_supplies_needed_to_make_product = all_pickle_path + "/supplies_needed_to_make_product_dict.pkl"

loc_of_cred_file = all_pickle_path + "/aws_cred.pkl"

REFRESH_TOKEN = ""
CLIENT_ID = ""
CLIENT_SECRET = ""
AWS_ACCESS_KEY = ""
AWS_SECRET_KEY = ""
ROLE_ARN = ""
MERCHANT_TOKEN = ""

access_token = ""

market_id_USA = "ATVPDKIKX0DER"
market_id_Canada = "A2EUQ1WTGCTBG2"
market_id_Mexico = "A1AM78C64UM0Y8"

loc_of_shipment_nums_excel_file = "shipment_nums.xlsx"
loc_of_check_excel_file = "shipment_nums_to_check.xlsx"

loc_of_shipment_info = "shipment_info"

long_days_ago = 32
short_term_days = 8
days_of_inventory_to_always_have_at_amazon = 60
shipment_frequency_days = 14

ShipFromAddress = {
    "addressLine1": "40828 FM149 RD",
    "city": "MAGNOLIA",
    "companyName": "Dlh Western LLC",
    "countryCode": "US",
    "email": "dlhwestern@gmail.com",
    "name": "Dlh Western",
    "phoneNumber": "9364198123",
    "postalCode": "77354-1225",
    "stateOrProvinceCode": "TX"
}

Contact_details_for_shipment = {
    "Email": "dlhwestern@gmail.com",
    "Name": "Dlh Western",
    "Phone": "9364198123"
}

parts = []
x_min = -1
x_max = -1
y_min = -1
y_max = -1



folder_for_cheapest_shipment = "data_for_finding_cheapest_shipment"
file_for_cheapest_shipment_info = "data.pkl"
file_for_cheapest_shipment_info_text = "cheapest.txt"
path_to_data_file_for_cheapest_shipment = folder_for_cheapest_shipment + "/" + file_for_cheapest_shipment_info
path_to_file_for_cheapest_shipment_info_text = folder_for_cheapest_shipment + "/" + file_for_cheapest_shipment_info_text



import firebase_admin
from firebase_admin import db

cred = firebase_admin.credentials.Certificate('products-in-bays-firebase-adminsdk-ut1ky-38a8b1fc77.json')
firebase_admin.initialize_app(cred, {'databaseURL': 'https://products-in-bays-default-rtdb.firebaseio.com/'})

sku_to_location_in_bays = {}


def visitor_body(text, cm, tm, fontDict, fontSize):
    x = tm[4]
    y = tm[5]
    if x > x_min and x < x_max and y > y_min and y < y_max and y > 4:  # had to had y > 4 to stop picking up all the carton nums on the bottom left corner label
        parts.append(text)


def product_label_pdf(our_sku_list, fnsku_list, title_list, condition_list, num_of_products_shipping_list,
                      file_path_name, shipment_id, list_of_items_per_carton):
    print("making product label pdf")

    pdfmetrics.registerFont(TTFont('consola', "CONSOLA.TTF"))

    factor_to_reduce_font = 0.8333333
    extra_labels_required = 5
    PAGESIZE = (612, 792)
    text_color = colors.black
    font_name = "consola"

    canvas_for_product_label_pdf = Canvas(file_path_name, pagesize=PAGESIZE)

    page_width = PAGESIZE[0]
    page_height = PAGESIZE[1]

    y_start_big_title_percent = 0.025
    x_start_big_title_percent = 0.035
    height_big_title_percent = 0.012
    width_big_title_percent = 1.0 - (2 * x_start_big_title_percent)

    y_start_shipment_id_percent = 0.04
    x_start_shipment_id_percent = 0.035
    height_shipment_id_percent = 0.012
    width_shipment_id_percent = 0.3

    y_start_items_per_carton_percent = 0.04
    x_start_items_per_carton_percent = 1 - 0.035
    height_items_per_carton_percent = 0.012
    width_items_per_carton_percent = x_start_items_per_carton_percent - (
            x_start_shipment_id_percent + width_shipment_id_percent)

    y_start_big_sku_percent = 0.047
    x_start_big_sku_percent = 0.035
    height_big_sku_percent = 0.06
    width_big_sku_percent = 1.0 - (2 * x_start_big_title_percent)

    y_label_start_percent = 0.045
    x_label_start_percent = 0.035
    x_label_space_percent = 0.014118
    y_label_space_percent = 0.0

    label_width_percent = 0.308823
    label_height_percent = 0.09090909

    y_barcode_start_from_label_corner_percent = 0.052272
    x_barcode_start_from_label_corner_percent = 0.0170588

    barcode_width_percent = 0.2270588
    barcode_height_percent = 0.019090909
    width_of_label_text_drawing_percent = 0.25882
    height_of_label_text_drawing_percent = 0.038636

    y_gap_between_barcode_and_text_drawing_percent = 0.0028
    # stuff in label_text_drawing
    height_of_label_fnsku_text_percent = 0.32
    width_of_label_fnsku_text_percent = 0.72
    y_label_fnsku_text_percent = 0.66
    x_label_fnsku_text_percent = 0.2727
    height_of_label_title_text_percent = 0.27
    width_of_label_title_text_percent = 1.0
    y_label_title_text_percent = 0.3
    x_label_title_text_percent = 0.0
    height_of_label_condition_text_percent = 0.28
    width_of_label_condition_text_percent = 1.0
    y_label_condition_text_percent = 0.0
    x_label_condition_text_percent = 0.0

    # actual coordinate vals

    y_start_big_title = y_start_big_title_percent * page_height
    x_start_big_title = x_start_big_title_percent * page_width
    height_big_title = height_big_title_percent * page_height
    width_big_title = width_big_title_percent * page_width

    y_start_big_sku = y_start_big_sku_percent * page_height
    x_start_big_sku = x_start_big_sku_percent * page_width
    height_big_sku = height_big_sku_percent * page_height
    width_big_sku = width_big_sku_percent * page_width

    y_start_shipment_id = y_start_shipment_id_percent * page_height
    x_start_shipment_id = x_start_shipment_id_percent * page_width
    height_shipment_id = height_shipment_id_percent * page_height
    width_shipment_id = width_shipment_id_percent * page_width

    y_start_items_per_carton = y_start_items_per_carton_percent * page_height
    x_start_items_per_carton = x_start_items_per_carton_percent * page_width
    height_items_per_carton = height_items_per_carton_percent * page_height
    width_items_per_carton = width_items_per_carton_percent * page_width

    y_label_start = y_label_start_percent * page_height
    x_label_start = x_label_start_percent * page_width
    x_label_space = x_label_space_percent * page_width
    y_label_space = y_label_space_percent * page_height

    label_width = label_width_percent * page_width
    label_height = label_height_percent * page_height

    y_barcode_start_from_label_corner = y_barcode_start_from_label_corner_percent * page_height
    x_barcode_start_from_label_corner = x_barcode_start_from_label_corner_percent * page_width

    barcode_width = barcode_width_percent * page_width
    barcode_height = barcode_height_percent * page_height
    width_of_label_text_drawing = width_of_label_text_drawing_percent * page_width
    height_of_label_text_drawing = height_of_label_text_drawing_percent * page_height

    y_gap_between_barcode_and_text_drawing = y_gap_between_barcode_and_text_drawing_percent * page_height
    # stuff in label_text_drawing
    height_of_label_fnsku_text = height_of_label_fnsku_text_percent * height_of_label_text_drawing
    width_of_label_fnsku_text = width_of_label_fnsku_text_percent * width_of_label_text_drawing
    y_label_fnsku_text = y_label_fnsku_text_percent * height_of_label_text_drawing
    x_label_fnsku_text = x_label_fnsku_text_percent * width_of_label_text_drawing
    height_of_label_title_text = height_of_label_title_text_percent * height_of_label_text_drawing
    width_of_label_title_text = width_of_label_title_text_percent * width_of_label_text_drawing
    y_label_title_text = y_label_title_text_percent * height_of_label_text_drawing
    x_label_title_text = x_label_title_text_percent * width_of_label_text_drawing
    height_of_label_condition_text = height_of_label_condition_text_percent * height_of_label_text_drawing
    width_of_label_condition_text = width_of_label_condition_text_percent * width_of_label_text_drawing
    y_label_condition_text = y_label_condition_text_percent * height_of_label_text_drawing
    x_label_condition_text = x_label_condition_text_percent * width_of_label_text_drawing

    for i in range(len(our_sku_list)):
        our_sku = our_sku_list[i]
        fnsku = fnsku_list[i]
        title = title_list[i]
        condition = condition_list[i]
        num_of_products_shipping = num_of_products_shipping_list[i]
        items_per_carton = list_of_items_per_carton[i]

        if len(fnsku) != 10:
            raise Exception('len(fnsku) != 10 for our_sku: ' + our_sku)

        total_labels_needed = num_of_products_shipping + extra_labels_required

        # make barcode
        barcode = Code128(fnsku)
        barcode._calculate()
        # the quiet space before and after the barcode
        quiet = barcode.lquiet + barcode.rquiet
        # total_wid = barWidth*charWid + quiet_space
        # char_wid = (total_width - quiet) / bar_width
        char_width = (barcode._width - quiet) / barcode.barWidth
        # now that we have the char width we can calculate the bar width
        bar_width = (barcode_width) / char_width
        # set the new bar width
        barcode.barWidth = bar_width
        barcode.barHeight = barcode_height
        # re-calculate
        barcode._calculate()
        barcode_x_offset = -1 * barcode.lquiet

        # test barcode
        canvas_for_barcode_test = Canvas("barcode_test.pdf", pagesize=(barcode_width * 1.2, barcode_height * 1.2))
        barcode.drawOn(canvas_for_barcode_test, 5 + barcode_x_offset, 0)
        canvas_for_barcode_test.save()
        images = convert_from_path("barcode_test.pdf")
        images[0].save("barcode_test.png", 'PNG')
        image = "barcode_test.png"
        img = cv2.imread(image)
        detectedBarcodes = decode(img)
        if not detectedBarcodes:
            raise Exception('barcode not detected for our_sku: ' + our_sku)
        else:
            if len(detectedBarcodes) > 1:
                raise Exception('len(detectedBarcodes) > 1 for our_sku: ' + our_sku)
            detect_barcode = detectedBarcodes[0]

            barcode_text = ""
            barcode_type = ""
            if detect_barcode.data != "":
                barcode_text = detect_barcode.data.decode("utf-8")
                barcode_type = detect_barcode.type
            if barcode_text != fnsku:
                raise Exception('barcode_text != fnsku for our_sku: ' + our_sku)
            if barcode_type != 'CODE128':
                raise Exception("barcode_type != 'CODE128' for sku: " + our_sku)
        os.remove("barcode_test.pdf")
        os.remove("barcode_test.png")

        # make drawing for label text
        drawing_for_label_text = Drawing(width_of_label_text_drawing, height_of_label_text_drawing)
        fnsku_text = String(x_label_fnsku_text,
                            y_label_fnsku_text, fnsku,
                            fontName=font_name,
                            fontSize=height_of_label_fnsku_text * factor_to_reduce_font,
                            fillColor=text_color)
        bounds = fnsku_text.getBounds()
        actual_width = bounds[2] - bounds[0]
        if actual_width > width_of_label_fnsku_text:
            raise Exception('actual_width > width_of_label_fnsku_text for our_sku: ' + our_sku)
        drawing_for_label_text.add(fnsku_text)

        title_text = String(x_label_title_text,
                            y_label_title_text, title,
                            fontName=font_name,
                            fontSize=height_of_label_title_text * factor_to_reduce_font,
                            fillColor=text_color)
        bounds = title_text.getBounds()
        actual_width = bounds[2] - bounds[0]
        if actual_width > width_of_label_title_text:
            # reduce title text
            last_char_change = 0
            last_tot_title_char_len = len(title)
            while True:
                if actual_width <= width_of_label_title_text and abs(last_char_change) <= 2:
                    break
                percent_diff = -1 * (actual_width - width_of_label_title_text) / actual_width
                last_char_change = round(percent_diff * last_tot_title_char_len)
                if percent_diff < 0 and last_char_change == 0:
                    last_char_change = -1
                last_tot_title_char_len = last_tot_title_char_len + last_char_change
                part_1_len = round(last_tot_title_char_len / 2)
                part_2_len = last_tot_title_char_len - part_1_len
                if part_1_len + part_2_len > len(title):
                    raise Exception('part_1_len+part_2_len > len(title) in label section for title: ' + title)
                new_title = title[0:part_1_len] + "..." + title[-part_2_len:]
                title_text = String(x_label_title_text,
                                    y_label_title_text, new_title,
                                    fontName=font_name,
                                    fontSize=height_of_label_title_text * factor_to_reduce_font,
                                    fillColor=text_color)

                bounds = title_text.getBounds()
                actual_width = bounds[2] - bounds[0]
        drawing_for_label_text.add(title_text)

        condition_text = String(x_label_condition_text,
                                y_label_condition_text, condition,
                                fontName=font_name,
                                fontSize=height_of_label_condition_text * factor_to_reduce_font,
                                fillColor=text_color)
        bounds = condition_text.getBounds()
        actual_width = bounds[2] - bounds[0]
        if actual_width > width_of_label_condition_text:
            raise Exception('actual_width > width_of_label_condition_text for our_sku: ' + our_sku)
        drawing_for_label_text.add(condition_text)

        labels_made = 0

        while labels_made < total_labels_needed:

            # add top of page stuff
            drawing_for_top_page_stuff = Drawing(page_width, y_label_start + label_height)
            max_width_for_big_title = width_big_title
            big_title_text = String(x_start_big_title,
                                    drawing_for_top_page_stuff.height - (y_start_big_title + height_big_title), title,
                                    fontName=font_name,
                                    fontSize=height_big_title * factor_to_reduce_font,
                                    fillColor=text_color)

            bounds = big_title_text.getBounds()
            actual_width = bounds[2] - bounds[0]
            if actual_width > max_width_for_big_title:
                # reduce text length
                last_char_change = 0
                last_tot_title_char_len = len(title)
                while True:
                    if actual_width <= max_width_for_big_title and abs(last_char_change) <= 2:
                        break
                    percent_diff = -1 * (actual_width - max_width_for_big_title) / actual_width
                    last_char_change = round(percent_diff * last_tot_title_char_len)
                    if percent_diff < 0 and last_char_change == 0:
                        last_char_change = -1
                    last_tot_title_char_len = last_tot_title_char_len + last_char_change
                    part_1_len = round(last_tot_title_char_len / 2)
                    part_2_len = last_tot_title_char_len - part_1_len
                    if part_1_len + part_2_len > len(title):
                        raise Exception('part_1_len+part_2_len > len(title) for title: ' + title)
                    new_title = title[0:part_1_len] + "..." + title[-part_2_len:]
                    big_title_text = String(x_start_big_title,
                                            drawing_for_top_page_stuff.height - (y_start_big_title + height_big_title),
                                            new_title,
                                            fontName=font_name,
                                            fontSize=height_big_title * factor_to_reduce_font,
                                            fillColor=text_color)

                    bounds = big_title_text.getBounds()
                    actual_width = bounds[2] - bounds[0]
            drawing_for_top_page_stuff.add(big_title_text)

            max_width_for_shipment_id = width_shipment_id
            shipment_id_text = String(x_start_shipment_id,
                                      drawing_for_top_page_stuff.height - (y_start_shipment_id + height_shipment_id),
                                      shipment_id,
                                      fontName=font_name,
                                      fontSize=height_shipment_id * factor_to_reduce_font,
                                      fillColor=text_color
                                      )
            bounds = shipment_id_text.getBounds()
            actual_width = bounds[2] - bounds[0]
            if actual_width > max_width_for_shipment_id:
                raise Exception('actual_width > max_width_for_shipment_id for sku:' + our_sku)
            drawing_for_top_page_stuff.add(shipment_id_text)

            max_width_for_items_per_carton = width_items_per_carton
            items_per_carton_text = String(x_start_items_per_carton,
                                           drawing_for_top_page_stuff.height - (
                                                   y_start_items_per_carton + height_items_per_carton),
                                           str(items_per_carton) + " items per carton",
                                           fontName=font_name,
                                           fontSize=height_items_per_carton * factor_to_reduce_font,
                                           fillColor=text_color,
                                           textAnchor="end"
                                           )
            bounds = items_per_carton_text.getBounds()
            actual_width = bounds[2] - bounds[0]
            if actual_width > max_width_for_items_per_carton:
                raise Exception('actual_width > max_width_for_items_per_carton for sku:' + our_sku)
            drawing_for_top_page_stuff.add(items_per_carton_text)

            max_width_for_big_sku = width_big_sku
            big_sku_text = String(x_start_big_sku,
                                  drawing_for_top_page_stuff.height - (y_start_big_sku + height_big_sku),
                                  "SKU " + our_sku,
                                  fontName=font_name,
                                  fontSize=height_big_sku * factor_to_reduce_font,
                                  fillColor=text_color
                                  )
            bounds = big_sku_text.getBounds()
            actual_width = bounds[2] - bounds[0]
            if actual_width > max_width_for_big_sku:
                raise Exception('actual_width > max_width_for_big_sku for sku:' + our_sku)
            drawing_for_top_page_stuff.add(big_sku_text)
            renderPDF.draw(drawing_for_top_page_stuff, canvas_for_product_label_pdf, 0,
                           page_height - drawing_for_top_page_stuff.height)

            # add labels
            for col_num in range(3):
                big_break = False
                left_x = x_label_start + (col_num * (label_width + x_label_space))
                for row_num in range(9):

                    bottom_y = page_height - (y_label_start + label_height + y_label_space + (
                            (row_num + 1) * (label_height + y_label_space)))

                    bar_x = left_x + x_barcode_start_from_label_corner + barcode_x_offset
                    bar_y = bottom_y + y_barcode_start_from_label_corner

                    barcode.drawOn(canvas_for_product_label_pdf, bar_x, bar_y)

                    text_drawing_x = left_x + x_barcode_start_from_label_corner
                    text_drawing_y = bar_y - (y_gap_between_barcode_and_text_drawing + height_of_label_text_drawing)
                    renderPDF.draw(drawing_for_label_text, canvas_for_product_label_pdf, text_drawing_x, text_drawing_y)

                    # canvas_for_product_label_pdf.setStrokeColorRGB(0.7,0.2,0.2)
                    # canvas_for_product_label_pdf.rect(left_x,bottom_y,label_width,label_height,1)

                    labels_made += 1
                    if labels_made >= total_labels_needed:
                        big_break = True
                        break
                if big_break == True:
                    break

            canvas_for_product_label_pdf.showPage()

    canvas_for_product_label_pdf.save()

    print("verifying product label pdf")
    print("")

    # check barcode readablilty and SKU correctness at top of page
    directory_name = "product_label_images_from_pdf"
    if os.path.exists(directory_name):
        shutil.rmtree(directory_name)
    os.mkdir(directory_name)
    images = convert_from_path(file_path_name)
    count = 0
    for i in images:
        i.save(directory_name + "/" + str(count) + ".png", 'PNG')
        count += 1
        message = "made image page " + str(count) + " out of " + str(len(images))
        sys.stdout.write('\r' + message)

    cur_image_num = 0

    print("")

    text_reader = easyocr.Reader(['en'])

    for i in range(len(our_sku_list)):
        our_sku = our_sku_list[i]
        fnsku = fnsku_list[i]
        num_of_products_shipping = num_of_products_shipping_list[i]

        total_labels_needed = num_of_products_shipping + extra_labels_required

        label_count = 0

        while label_count < total_labels_needed:
            message = "verified image page " + str(cur_image_num + 1) + " out of " + str(count)
            sys.stdout.write('\r' + message)
            image_name = directory_name + "/" + str(cur_image_num) + ".png"
            img_of_page = cv2.imread(image_name)
            img_height = int(img_of_page.shape[0])
            img_width = int(img_of_page.shape[1])

            # check big sku
            y_start_big_sku = round(y_start_big_sku_percent * img_height) - 1
            x_start_big_sku = round(x_start_big_sku_percent * img_width) - 1
            height_big_sku = round(height_big_sku_percent * img_height) + 2
            width_big_sku = round(width_big_sku_percent * img_width) + 2

            big_sku_text_img = img_of_page[y_start_big_sku:y_start_big_sku + height_big_sku,
                               x_start_big_sku:x_start_big_sku + width_big_sku]
            big_sku_text_img = cv2.resize(big_sku_text_img, (0, 0), fx=0.2, fy=0.2)
            big_sku_text_img = cv2.copyMakeBorder(big_sku_text_img, 60, 60, 60, 60,
                                                  cv2.BORDER_CONSTANT, None, [255, 255, 255])

            # cv2.imshow("cropped", big_sku_text_img)
            # cv2.waitKey(0)
            # x = input()

            # old way to check sku text

            # gray = cv2.cvtColor(big_sku_text_img, cv2.COLOR_BGR2GRAY)
            # ret, thresh1 = cv2.threshold(gray, 0, 255, cv2.THRESH_OTSU | cv2.THRESH_BINARY_INV)
            # rect_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (18, 18))
            # dilation = cv2.dilate(thresh1, rect_kernel, iterations=1)
            # contours, hierarchy = cv2.findContours(dilation, cv2.RETR_EXTERNAL,
            #                                        cv2.CHAIN_APPROX_NONE)
            # cnt = contours[0]
            # x, y, w, h = cv2.boundingRect(cnt)
            # # Cropping the text block for giving input to OCR
            # cropped = big_sku_text_img[y:y + h, x:x + w]
            # # Apply OCR on the cropped image
            # text = str(pytesseract.image_to_string(cropped)).replace("\n", "")
            # if text != "SKU " + our_sku:
            #     raise Exception(
            #         'text != our_sku for sku:' + our_sku + "  text: " + text + "  on page: " + str(cur_image_num))

            results = text_reader.readtext(big_sku_text_img, paragraph=True)
            text = results[0][1]
            if text != "SKU " + our_sku:
                if text != "SKU 46-7 --5":    #here we enter bugs in the reader program
                    raise Exception(
                        'text != our_sku for sku:' + our_sku + "  text: " + text + "  on page: " + str(cur_image_num))

            x_label_start_image = x_label_start_percent * img_width
            label_width_image = label_width_percent * img_width
            x_label_space_image = x_label_space_percent * img_width

            y_label_start_image = y_label_start_percent * img_height
            label_height_image = label_height_percent * img_height
            y_label_space_image = y_label_space_percent * img_height

            # check every barcode
            for col_num in range(3):
                big_break = False
                left_x = round(x_label_start_image + (col_num * (label_width_image + x_label_space_image)))
                right_x = round(left_x + label_width_image)
                for row_num in range(9):
                    bottom_y = round(y_label_start_image + label_height_image + y_label_space_image + (
                            (row_num + 1) * (label_height_image + y_label_space_image)))
                    top_y = round(bottom_y - label_height_image)

                    # print("col_num: "+str(col_num)+"  row_num: "+str(row_num))
                    # print("left_x: "+str(left_x))
                    # print("right_x: "+str(right_x))
                    # print("bottom_y: "+str(bottom_y))
                    # print("top_y: "+str(top_y))
                    # print()

                    label_img = img_of_page[top_y:bottom_y, left_x:right_x]

                    detectedBarcodes = decode(label_img)
                    if not detectedBarcodes:
                        raise Exception('barcode not detected during label testing for our_sku: ' + our_sku)
                    else:
                        if len(detectedBarcodes) > 1:
                            raise Exception('len(detectedBarcodes) > 1 during label testing for our_sku: ' + our_sku)
                        detect_barcode = detectedBarcodes[0]

                        barcode_text = ""
                        barcode_type = ""
                        if detect_barcode.data != "":
                            barcode_text = detect_barcode.data.decode("utf-8")
                            barcode_type = detect_barcode.type
                        if barcode_text != fnsku:
                            raise Exception('barcode_text != fnsku during label testing for our_sku: ' + our_sku)
                        if barcode_type != 'CODE128':
                            raise Exception("barcode_type != 'CODE128' during label testing for sku: " + our_sku)

                    label_count += 1
                    if label_count >= total_labels_needed:
                        big_break = True
                        break
                if big_break == True:
                    break

            cur_image_num += 1

    # delete everything
    for i in range(count):
        os.remove(directory_name + "/" + str(i) + ".png")
    os.rmdir(directory_name)

    print("")
    print("product labels done for shipment: " + str(shipment_id))


def add_case_names_to_carton_labels(padded_carton_num_to_case_name, shipment_id, package_label_path,
                                    package_label_with_case_name_path, case_name_to_amazon_sku):
    global parts
    global x_min
    global x_max
    global y_min
    global y_max

    existing_pdf = PdfReader(open(package_label_path, "rb"))
    output = PdfWriter()

    list_of_padded_carton_num = list(padded_carton_num_to_case_name.keys())
    list_of_padded_carton_num.sort()

    spot_in_list = -1

    num_of_pages = len(existing_pdf.pages)

    og_width = 0
    og_height = 0

    for i in range(num_of_pages):
        page = existing_pdf.pages[i]
        packet = io.BytesIO()
        can = canvas.Canvas(packet)

        mediabox = page.mediabox
        height = mediabox.height
        width = mediabox.width

        font_height = round(0.015 * height)

        can.setPageSize((width, height))
        can.setFont("Helvetica", font_height)

        if i == 0:
            og_width = width
            og_height = height

        if og_width != width or og_height != height:
            raise Exception('og_width != width or og_height != height for shipment_id: ' + shipment_id)

        third_of_height = round(height / 3)
        two_thirds_of_height = round(2 * height / 3)
        half_of_width = round(width / 2)

        # lowest_height_per = 0.071969
        lowest_height_per = 0.08
        lowest_height = round(lowest_height_per * height)

        height_spacing_per = .3035
        height_spacing = round(height_spacing_per * height)

        lowest_width_per = 0.037
        lowest_width = round(lowest_width_per * width)

        width_spacing_per = .488
        width_spacing = round(width_spacing_per * width)

        for j in range(6):
            spot_in_list += 1

            x_of_text = 0
            y_of_text = 0

            if j == 0:
                x_min = 0
                x_max = half_of_width
                y_min = two_thirds_of_height
                y_max = height

                x_of_text = lowest_width
                y_of_text = lowest_height + (2 * height_spacing)
            if j == 1:
                x_min = 0
                x_max = half_of_width
                y_min = third_of_height
                y_max = two_thirds_of_height

                x_of_text = lowest_width
                y_of_text = lowest_height + height_spacing
            if j == 2:
                x_min = 0
                x_max = half_of_width
                y_min = 0
                y_max = third_of_height

                x_of_text = lowest_width
                y_of_text = lowest_height
            if j == 3:
                x_min = half_of_width
                x_max = width
                y_min = two_thirds_of_height
                y_max = height

                x_of_text = lowest_width + width_spacing
                y_of_text = lowest_height + (2 * height_spacing)
            if j == 4:
                x_min = half_of_width
                x_max = width
                y_min = third_of_height
                y_max = two_thirds_of_height

                x_of_text = lowest_width + width_spacing
                y_of_text = lowest_height + height_spacing
            if j == 5:
                x_min = half_of_width
                x_max = width
                y_min = 0
                y_max = third_of_height

                x_of_text = lowest_width + width_spacing
                y_of_text = lowest_height

            parts = []
            page.extract_text(visitor_text=visitor_body)
            text_body = "".join(parts)
            count_of_carton_num = text_body.count(shipment_id + "U")
            if count_of_carton_num != 1:
                raise Exception('count_of_carton_num != 1 for shipment_id: ' + shipment_id)

            cur_carton_id = shipment_id + list_of_padded_carton_num[spot_in_list]
            if cur_carton_id not in text_body:
                raise Exception('cur_carton_id not in text_body for shipment_id: ' + shipment_id)

            case_name = padded_carton_num_to_case_name[list_of_padded_carton_num[spot_in_list]]

            can.drawString(x_of_text, y_of_text, case_name)

            if spot_in_list == len(list_of_padded_carton_num) - 1:
                break

        can.save()

        # move to the beginning of the StringIO buffer
        packet.seek(0)

        # create a new PDF with Reportlab
        new_pdf = PdfReader(packet)

        page.merge_page(new_pdf.pages[0])
        output.add_page(page)

        if i == (num_of_pages - 1):
            if spot_in_list != len(list_of_padded_carton_num) - 1:
                raise Exception('spot_in_list != len(list_of_padded_carton_num)-1 for shipment_id: ' + shipment_id)

    output_stream = open(package_label_with_case_name_path, "wb")
    output.write(output_stream)
    output_stream.close()

    # read new pdf and double check right case name is on right carton id

    new_pdf = PdfReader(open(package_label_with_case_name_path, "rb"))
    spot_in_list = -1

    num_of_pages = len(new_pdf.pages)

    og_width = 0
    og_height = 0

    for i in range(num_of_pages):
        page = new_pdf.pages[i]

        mediabox = page.mediabox
        height = mediabox.height
        width = mediabox.width

        if i == 0:
            og_width = width
            og_height = height

        if og_width != width or og_height != height:
            raise Exception('og_width != width or og_height != height for shipment_id: ' + shipment_id)

        third_of_height = round(height / 3)
        two_thirds_of_height = round(2 * height / 3)
        half_of_width = round(width / 2)

        y_offset = -20

        for j in range(6):
            spot_in_list += 1

            if j == 0:
                x_min = 0
                x_max = half_of_width
                y_min = two_thirds_of_height + y_offset
                y_max = height + y_offset
            if j == 1:
                x_min = 0
                x_max = half_of_width
                y_min = third_of_height + y_offset
                y_max = two_thirds_of_height + y_offset
            if j == 2:
                x_min = 0
                x_max = half_of_width
                y_min = 0 + y_offset
                y_max = third_of_height + y_offset
            if j == 3:
                x_min = half_of_width
                x_max = width
                y_min = two_thirds_of_height + y_offset
                y_max = height + y_offset
            if j == 4:
                x_min = half_of_width
                x_max = width
                y_min = third_of_height + y_offset
                y_max = two_thirds_of_height + y_offset
            if j == 5:
                x_min = half_of_width
                x_max = width
                y_min = 0 + y_offset
                y_max = third_of_height + y_offset

            parts = []
            page.extract_text(visitor_text=visitor_body)
            text_body = "".join(parts)
            count_of_carton_num = text_body.count(shipment_id + "U")
            if count_of_carton_num != 1:
                raise Exception('count_of_carton_num != 1 for shipment_id: ' + shipment_id)

            cur_carton_id = shipment_id + list_of_padded_carton_num[spot_in_list]
            if cur_carton_id not in text_body:
                raise Exception('cur_carton_id not in text_body for shipment_id: ' + shipment_id)

            case_name = padded_carton_num_to_case_name[list_of_padded_carton_num[spot_in_list]]

            if case_name not in text_body:
                raise Exception('case_name not in text_body for shipment_id: ' + shipment_id)

            should_be_amazon_sku = case_name_to_amazon_sku[case_name]
            first_part_of_should_be_sku = should_be_amazon_sku[:9]
            last_part_of_should_be_sku = should_be_amazon_sku[-8:]


            if first_part_of_should_be_sku not in text_body or last_part_of_should_be_sku not in text_body:
                raise Exception('should_be_amazon_sku not in text_body for shipment_id: ' + shipment_id)

            if spot_in_list == len(list_of_padded_carton_num) - 1:
                break

        if i == (num_of_pages - 1):
            if spot_in_list != len(list_of_padded_carton_num) - 1:
                raise Exception('spot_in_list != len(list_of_padded_carton_num)-1 for shipment_id: ' + shipment_id)


def prettify(elem):
    """Return a pretty-printed XML string for the Element.
    """
    rough_string = ET.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="\t")


def get_aws_cred():
    global REFRESH_TOKEN
    global CLIENT_ID
    global CLIENT_SECRET
    global AWS_ACCESS_KEY
    global AWS_SECRET_KEY
    global ROLE_ARN
    global MERCHANT_TOKEN

    cred_dict = pickle.load(open(loc_of_cred_file, "rb"))
    REFRESH_TOKEN = cred_dict["REFRESH_TOKEN"]
    CLIENT_ID = cred_dict["CLIENT_ID"]
    CLIENT_SECRET = cred_dict["CLIENT_SECRET"]
    AWS_ACCESS_KEY = cred_dict["AWS_ACCESS_KEY"]
    AWS_SECRET_KEY = cred_dict["AWS_SECRET_KEY"]
    ROLE_ARN = cred_dict["ROLE_ARN"]
    MERCHANT_TOKEN = cred_dict["MERCHANT_TOKEN"]


def get_access_token():
    global access_token
    url = "https://api.amazon.com/auth/o2/token"

    payload = 'grant_type=refresh_token&refresh_token=' + REFRESH_TOKEN + '&client_id=' + CLIENT_ID + '&client_secret=' + CLIENT_SECRET
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    response = requests.request("POST", url, headers=headers, data=payload).json()
    access_token = response["access_token"]


def before_run():
    all_files = os.listdir(all_pickle_path)
    try:
        for name in all_files:
            if name != ".DS_Store":
                test_read = pickle.load(open(all_pickle_path + "/" + name, "rb"))
    except:
        print("Error: pickle files corrupted before run")
        exit()

    if os.path.exists(all_pickle_path_copy):
        shutil.rmtree(all_pickle_path_copy)
    copy_tree(all_pickle_path, all_pickle_path_copy)

    all_files = os.listdir(all_pickle_path_copy)
    try:
        for name in all_files:
            if name != ".DS_Store":
                test_read = pickle.load(open(all_pickle_path_copy + "/" + name, "rb"))
    except:
        print("Error: copied pickle files corrupted before run")
        exit()


def after_run():
    all_files = os.listdir(all_pickle_path)
    try:
        for name in all_files:
            if name != ".DS_Store":
                test_read = pickle.load(open(all_pickle_path + "/" + name, "rb"))
    except:
        print("Error: pickle files corrupted after run")
        exit()


def make_excel_temp():
    if os.path.exists(loc_of_shipment_nums_excel_file):
        os.remove(loc_of_shipment_nums_excel_file)

    if os.path.exists(loc_of_check_excel_file):
        os.remove(loc_of_check_excel_file)

    amazon_fba_sku_dict = {}
    if os.path.exists(loc_of_amazon_fba_sku_dict):
        amazon_fba_sku_dict = pickle.load(open(loc_of_amazon_fba_sku_dict, "rb"))

    amazon_sku_to_par_var_dict = pickle.load(open(loc_of_supply_name_file, "rb"))

    new_skus_need_data_for = []

    for sku in amazon_sku_to_par_var_dict:
        if sku not in amazon_fba_sku_dict:
            new_skus_need_data_for.append(sku)

    if len(new_skus_need_data_for) > 0:
        inv_list = get_inventory(new_skus_need_data_for)
        for data in inv_list:
            sku = data["sellerSku"]
            print("getting info for sku: " + sku)
            asin = data["asin"]
            product_name = data["productName"]
            creation_timestamp = get_creation_time_UTC(sku)
            dict = {}
            dict["asin"] = asin
            dict["product_name"] = product_name
            dict["creation_timestamp"] = creation_timestamp
            dict["if_mature"] = False
            amazon_fba_sku_dict[sku] = dict

    skus_of_excel = list(amazon_fba_sku_dict.keys())

    dict_of_warnings = {}  # sku to list of warnings
    for sku in skus_of_excel:
        dict_of_warnings[sku] = []

    inventory_count_of_sku = {}

    max_percent_off_from_amzn_inv = 0.2
    max_percent_of_researching_quantity = 0.2

    # skus_of_excel = ["HX-9YX3-ZMGH"]

    tot_data = get_inventory(skus_of_excel)
    for data in tot_data:
        sku = data["sellerSku"]
        inv_details = data["inventoryDetails"]
        good_inv = 0
        fufillable = int(inv_details["fulfillableQuantity"])
        good_inv = good_inv + fufillable
        tot_reserved = int(inv_details["reservedQuantity"]["totalReservedQuantity"])
        customer_order = int(inv_details["reservedQuantity"]["pendingCustomerOrderQuantity"])
        good_inv = good_inv + (tot_reserved - customer_order)

        inbound_shipped = int(inv_details["inboundShippedQuantity"])
        inbound_receiving = int(inv_details["inboundReceivingQuantity"])
        inbound_working = int(inv_details["inboundWorkingQuantity"])
        good_inv = good_inv + inbound_shipped + inbound_receiving + inbound_working

        researching_quantity = 0
        if "researchingQuantity" in inv_details:
            researching_quantity = int(inv_details["researchingQuantity"]["totalResearchingQuantity"])

        amzn_tot_inv = int(data["totalQuantity"])

        if amzn_tot_inv == 0:
            inventory_count_of_sku[sku] = 0
        else:
            per_diff = abs(good_inv - amzn_tot_inv) / amzn_tot_inv
            if per_diff > max_percent_off_from_amzn_inv:
                tot_unfilfilable = int(inv_details["unfulfillableQuantity"]["totalUnfulfillableQuantity"])
                if abs(tot_unfilfilable - amzn_tot_inv) < 5:
                    # this is ok, cause the issue is coming from unfulfilable
                    x = 0
                else:
                    dict_of_warnings[sku].append("good inventory different from total inventory")
            if researching_quantity >= (good_inv * max_percent_of_researching_quantity):
                dict_of_warnings[sku].append("large amount of researching quantity")

            inventory_count_of_sku[sku] = good_inv

    # get sales history
    total_days_of_inventory_to_have = days_of_inventory_to_always_have_at_amazon + shipment_frequency_days

    raw_sales_data = sales_for_skus(skus_of_excel, long_days_ago)
    # raw_sales_data = pickle.load(open("raw_sales.pkl", "rb"))

    inventory_level_per_sku = {}
    for sku in raw_sales_data:
        cur_list = raw_sales_data[sku]  # spot 0 is oldest spot
        long_tot = 0
        short_tot = 0
        for i in range(len(cur_list) - 1, 0, -1):
            val = cur_list[i]
            long_tot += val
            if (len(cur_list) - 1) - i < short_term_days:
                short_tot += val
        extrapolated_short_tot = float(long_days_ago / short_term_days) * short_tot
        avg_tot = float(long_tot + extrapolated_short_tot) / 2
        needed_inventory = round(float(total_days_of_inventory_to_have / long_days_ago) * avg_tot)
        dict = {}
        dict["short_term_sales"] = short_tot
        dict["long_term_sales"] = long_tot
        dict["total_inventory_required"] = needed_inventory
        need_to_ship = needed_inventory - inventory_count_of_sku[sku]
        if need_to_ship < 0:
            need_to_ship = 0
        dict["need_to_ship"] = need_to_ship
        inventory_level_per_sku[sku] = dict

    # making sure 2" foam is good

    two_inch_2_peice_amazon_sku = ""
    two_inch_4_peice_amazon_sku = ""
    for amazon_sku in amazon_sku_to_par_var_dict:
        par_var = amazon_sku_to_par_var_dict[amazon_sku]
        if par_var == "118-1":
            two_inch_2_peice_amazon_sku = amazon_sku
        if par_var == "118-2":
            two_inch_4_peice_amazon_sku = amazon_sku

    two_inch_2_peice_ship = 0
    if two_inch_2_peice_amazon_sku in inventory_level_per_sku:
        two_inch_2_peice_ship = inventory_level_per_sku[two_inch_2_peice_amazon_sku]["need_to_ship"]

    two_inch_4_peice_ship = 0
    if two_inch_4_peice_amazon_sku in inventory_level_per_sku:
        two_inch_4_peice_ship = inventory_level_per_sku[two_inch_4_peice_amazon_sku]["need_to_ship"]

    four_piece_equalvant = two_inch_4_peice_ship * 2

    if four_piece_equalvant > two_inch_2_peice_ship:
        inventory_level_per_sku[two_inch_2_peice_amazon_sku]["need_to_ship"] = four_piece_equalvant

    if four_piece_equalvant < two_inch_2_peice_ship:
        inventory_level_per_sku[two_inch_4_peice_amazon_sku]["need_to_ship"] = m.ceil(two_inch_2_peice_ship / 2)

    order_of_skus_for_excel = []  # spot 0 is newest
    order_of_creation_time = []
    for sku in skus_of_excel:
        creation_time = amazon_fba_sku_dict[sku]["creation_timestamp"]
        if_added = False
        for i in range(len(order_of_skus_for_excel)):
            check_time = order_of_creation_time[i]
            if creation_time > check_time:
                order_of_skus_for_excel.insert(i, sku)
                order_of_creation_time.insert(i, creation_time)
                if_added = True
                break
        if if_added == False:
            order_of_skus_for_excel.append(sku)
            order_of_creation_time.append(creation_time)

    # make excel

    workbook = openpyxl.Workbook()
    view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
    workbook.views = view
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    sheet.protection.sheet = True

    yellow_fill = PatternFill(start_color='FFFFFF00',
                              end_color='FFFFFF00',
                              fill_type='solid')
    orange_fill = PatternFill(start_color='FFF28C28',
                              end_color='FFF28C28',
                              fill_type='solid')

    no_fill = PatternFill(start_color='00FFFFFF',
                          end_color='00FFFFFF',
                          fill_type='solid')

    sheet["A1"] = "not mature"
    sheet["B1"] = "warning"
    sheet['A1'].alignment = Alignment(wrap_text=True)
    sheet['B1'].alignment = Alignment(wrap_text=True)
    sheet['A1'].fill = yellow_fill
    sheet['B1'].fill = orange_fill
    sheet.row_dimensions[1].height = 32
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 90
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 11
    sheet.column_dimensions['F'].width = 11
    sheet.column_dimensions['G'].width = 11
    sheet.column_dimensions['H'].width = 11
    sheet.column_dimensions['J'].width = 100

    add_info_row(3, sheet)

    cur_row = 5
    for sku in order_of_skus_for_excel:
        product_name = amazon_fba_sku_dict[sku]["product_name"]
        short_term_sales = inventory_level_per_sku[sku]["short_term_sales"]
        long_term_sales = inventory_level_per_sku[sku]["long_term_sales"]
        total_inventory_required = inventory_level_per_sku[sku]["total_inventory_required"]
        current_inv = inventory_count_of_sku[sku]
        need_to_ship = inventory_level_per_sku[sku]["need_to_ship"]
        if_mature = amazon_fba_sku_dict[sku]["if_mature"]
        warnings = []
        if sku in dict_of_warnings:
            warnings = dict_of_warnings[sku]

        sheet["A" + str(cur_row)] = amazon_sku_to_par_var_dict[sku]
        sheet["B" + str(cur_row)] = sku
        sheet["C" + str(cur_row)] = product_name
        sheet["D" + str(cur_row)] = short_term_sales
        sheet["E" + str(cur_row)] = long_term_sales
        sheet["F" + str(cur_row)] = total_inventory_required
        sheet["G" + str(cur_row)] = current_inv
        sheet["H" + str(cur_row)] = need_to_ship
        sheet["I" + str(cur_row)] = if_mature
        sheet["J" + str(cur_row)] = str(warnings)
        if len(warnings) == 0:
            sheet["J" + str(cur_row)] = ""

        if if_mature == False:
            for cell in sheet[str(cur_row) + ":" + str(cur_row)]:
                cell.fill = yellow_fill
        if len(warnings) > 0:
            for cell in sheet[str(cur_row) + ":" + str(cur_row)]:
                cell.fill = orange_fill

        sheet["H" + str(cur_row)].fill = no_fill
        sheet["I" + str(cur_row)].fill = no_fill
        sheet["H" + str(cur_row)].protection = Protection(locked=False)
        sheet["I" + str(cur_row)].protection = Protection(locked=False)

        cur_row += 1

        if cur_row % 35 == 0:
            cur_row += 1
            add_info_row(cur_row, sheet)
            cur_row = cur_row + 2

    workbook.save(loc_of_shipment_nums_excel_file)

    pickle.dump(amazon_fba_sku_dict, open(loc_of_amazon_fba_sku_dict, "wb"))


def add_info_row(row_num, sheet):
    sheet["A" + str(row_num)] = "Sku"
    sheet["A" + str(row_num)].font = Font(bold=True)
    sheet["A" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["B" + str(row_num)] = "Amazon Sku"
    sheet["B" + str(row_num)].font = Font(bold=True)
    sheet["B" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["C" + str(row_num)] = "Product Name"
    sheet["C" + str(row_num)].font = Font(bold=True)
    sheet["C" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["D" + str(row_num)] = str(short_term_days) + " day sales"
    sheet["D" + str(row_num)].font = Font(bold=True)
    sheet["D" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["E" + str(row_num)] = str(long_days_ago) + " day sales"
    sheet["E" + str(row_num)].font = Font(bold=True)
    sheet["E" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["F" + str(row_num)] = "Inventory Required"
    sheet["F" + str(row_num)].font = Font(bold=True)
    sheet["F" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["G" + str(row_num)] = "Current Inventory"
    sheet["G" + str(row_num)].font = Font(bold=True)
    sheet["G" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["H" + str(row_num)] = "Shipment Amount"
    sheet["H" + str(row_num)].font = Font(bold=True)
    sheet["H" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["I" + str(row_num)] = "If Mature"
    sheet["I" + str(row_num)].font = Font(bold=True)
    sheet["I" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["J" + str(row_num)] = "Warnings"
    sheet["J" + str(row_num)].font = Font(bold=True)
    sheet["J" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[row_num].height = 30


def add_info_row_for_check_file(row_num, sheet):
    sheet["A" + str(row_num)] = "Sku"
    sheet["A" + str(row_num)].font = Font(bold=True)
    sheet["A" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["B" + str(row_num)] = "Amazon Sku"
    sheet["B" + str(row_num)].font = Font(bold=True)
    sheet["B" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["C" + str(row_num)] = "Product Name"
    sheet["C" + str(row_num)].font = Font(bold=True)
    sheet["C" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["D" + str(row_num)] = "Wanted Shipment Amount"
    sheet["D" + str(row_num)].font = Font(bold=True)
    sheet["D" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["E" + str(row_num)] = "Actual Shipment Amount"
    sheet["E" + str(row_num)].font = Font(bold=True)
    sheet["E" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet["F" + str(row_num)] = "Supplies Lacking"
    sheet["F" + str(row_num)].font = Font(bold=True)
    sheet["F" + str(row_num)].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[row_num].height = 30


def sales_for_skus(skus, days_ago):
    return_data = {}

    for sku in skus:
        end_date_obj = datetime.today()
        start_date_obj = datetime.today() - timedelta(days=days_ago)
        end_date_str = end_date_obj.strftime('%Y-%m-%d') + "T00:00:00-07:00"
        start_date_str = start_date_obj.strftime('%Y-%m-%d') + "T00:00:00-07:00"

        time_par = start_date_str + "--" + end_date_str

        response = {}

        while (True):
            url = "https://sellingpartnerapi-na.amazon.com/sales/v1/orderMetrics?marketplaceIds=" + market_id_USA + \
                  "&interval=" + time_par + "&granularity=Day&sku=" + sku
            payload = {}
            headers = {
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("GET", url, headers=headers, data=payload).json()

            need_to_pause = False
            if "errors" in response:
                str = response['errors'][0]["code"]
                if str == "QuotaExceeded":
                    need_to_pause = True
            if need_to_pause == True:
                print("quota hit, sleeping now")
                time.sleep(8)
            else:
                break

        list_of_days = response["payload"]
        list_of_units_sold_USA = []
        for data in list_of_days:
            units = int(data["unitCount"])
            list_of_units_sold_USA.append(units)

        # url = "https://sellingpartnerapi-na.amazon.com/sales/v1/orderMetrics?marketplaceIds=" + market_id_Canada + \
        #       "&interval=" + time_par + "&granularity=Day&sku=" + sku
        # payload = {}
        # headers = {
        #     'Accept': 'application/json',
        #     'x-amz-access-token': access_token
        # }
        # response = requests.request("GET", url, headers=headers, data=payload).json()
        # list_of_days = response["payload"]
        # list_of_units_sold_Canada = []
        # for data in list_of_days:
        #     units = int(data["unitCount"])
        #     list_of_units_sold_Canada.append(units)
        #
        #
        #
        # url = "https://sellingpartnerapi-na.amazon.com/sales/v1/orderMetrics?marketplaceIds=" + market_id_Mexico + \
        #       "&interval=" + time_par + "&granularity=Day&sku=" + sku
        # payload = {}
        # headers = {
        #     'Accept': 'application/json',
        #     'x-amz-access-token': access_token
        # }
        # response = requests.request("GET", url, headers=headers, data=payload).json()
        # list_of_days = response["payload"]
        # list_of_units_sold_Mexico = []
        # for data in list_of_days:
        #     units = int(data["unitCount"])
        #     list_of_units_sold_Mexico.append(units)

        total_units_sold = []
        for i in range(len(list_of_units_sold_USA)):
            # tot = list_of_units_sold_USA[i] + list_of_units_sold_Canada[i] + list_of_units_sold_Mexico[i]
            tot = list_of_units_sold_USA[i]
            total_units_sold.append(tot)

        return_data[sku] = total_units_sold
        print("got sales data for: " + sku)

    return return_data


def date_to_utc_time(str):
    utc_time = datetime.strptime(str, "%Y-%m-%dT%H:%M:%S.%fZ")
    milliseconds = (utc_time - datetime(1970, 1, 1)) // timedelta(milliseconds=1)
    return milliseconds


def get_creation_time_UTC(sku):
    time.sleep(0.22)

    url = "https://sellingpartnerapi-na.amazon.com/listings/2021-08-01/items/" + MERCHANT_TOKEN + "/" + sku + \
          "?marketplaceIds=" + market_id_USA + "&issueLocale=en_US&includedData=summaries"

    payload = {}
    headers = {
        'Accept': 'application/json',
        'x-amz-access-token': access_token
    }

    response = requests.request("GET", url, headers=headers, data=payload).json()
    str_time = response["summaries"][0]["createdDate"]
    milisec_UTC = date_to_utc_time(str_time)
    return milisec_UTC


def get_inventory(list_of_skus):
    max_sku_num_req = 49
    time_between_requests = 1

    big_list = []

    temp_list = []
    for i in range(len(list_of_skus)):
        sku = list_of_skus[i]
        temp_list.append(sku)
        if len(temp_list) >= max_sku_num_req or i == len(list_of_skus) - 1:
            str = ""
            for s in temp_list:
                str = str + s + ","
            str = str[:-1]

            url = "https://sellingpartnerapi-na.amazon.com/fba/inventory/v1/summaries?details=true&granularityType=Marketplace&granularityId=" + market_id_USA + "&sellerSkus=" + str + "&marketplaceIds=" + market_id_USA

            payload = {}
            headers = {
                'Accept': 'application/json',
                'x-amz-access-token': access_token,
                'X-Amz-Date': '20231002T210638Z',
                'Authorization': 'AWS4-HMAC-SHA256 Credential=AKIAWBVAKQIMKFKH77DB/20231002/us-east-1/execute-api/aws4_request, SignedHeaders=accept;host;x-amz-access-token;x-amz-date, Signature=6fd79f5b8fee02163561cf0d48dd2de9d879c15bb62277c9154affac190a461a'
            }

            response = requests.request("GET", url, headers=headers, data=payload).json()
            inv = response["payload"]["inventorySummaries"]
            big_list.extend(inv)

            temp_list.clear()
            time.sleep(time_between_requests)
    return big_list


def is_num_whole(num):
    clean = round(num, 7)
    if clean % 1 == 0:
        return True
    else:
        return False


def get_location_of_skus(amazon_sku_to_list_of_cases_wanting_to_ship_copy_for_func):
    global sku_to_location_in_bays
    skus_to_get = {}
    for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship_copy_for_func:
        list_of_case = amazon_sku_to_list_of_cases_wanting_to_ship_copy_for_func[amazon_sku]
        for dict in list_of_case:
            sku = dict['item_sku']
            skus_to_get[sku] = True

    for sku in skus_to_get:
        locs = db.reference('info/' + sku).get()
        if locs == None:
            sku_to_location_in_bays[sku] = ""
        else:
            locations = list(locs.keys())
            string_loc = ', '.join(locations)
            sku_to_location_in_bays[sku] = string_loc


def read_excel():
    while True:
        print('modify the shipment numbers in the "shipment_nums.xlsx" then enter "done"')
        x = input()
        if x == "done":
            break

    amazon_fba_sku_dict = pickle.load(open(loc_of_amazon_fba_sku_dict, "rb"))
    amazon_sku_to_par_var_dict = pickle.load(open(loc_of_supply_name_file, "rb"))
    workbook = openpyxl.load_workbook(loc_of_shipment_nums_excel_file)
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    cur_row = 1
    num_of_none = 0
    amazon_sku_to_wanted_shipment_amount = {}
    amazon_sku_to_title = {}
    order_of_skus_in_og_excel = []
    while True:
        name_val = sheet["B" + str(cur_row)].value
        if name_val == None:
            num_of_none += 1
        else:
            num_of_none = 0
            if name_val in amazon_fba_sku_dict:
                amazon_sku = name_val
                order_of_skus_in_og_excel.append(amazon_sku)
                amazon_sku_to_title[amazon_sku] = str(sheet["C" + str(cur_row)].value)
                string_mature = str(sheet["I" + str(cur_row)].value).lower()
                if string_mature == "true":
                    amazon_fba_sku_dict[amazon_sku]["if_mature"] = True
                elif string_mature == "false":
                    amazon_fba_sku_dict[amazon_sku]["if_mature"] = False
                else:
                    raise Exception("mature value is messed up for sku: " + amazon_sku)

                int_shipment_amount = int(sheet["H" + str(cur_row)].value)
                if int_shipment_amount < 0:
                    raise Exception("shipment amount is negative for sku: " + amazon_sku)

                if int_shipment_amount > 0:
                    amazon_sku_to_wanted_shipment_amount[amazon_sku] = int_shipment_amount

        if num_of_none > 10:
            break

        cur_row += 1

    sku_inventory_data = pickle.load(open(loc_of_sku_inventory_file, "rb"))

    test_supplies_levels = pickle.load(open(loc_of_supply_inventory_file, "rb"))

    supplies_needed_to_make_skus = pickle.load(open(loc_of_supplies_needed_to_make_product, "rb"))

    amazon_sku_to_list_of_cases_wanting_to_ship = {}

    amazon_skus_that_dont_have_enough_inventory = {}

    par_var_to_if_we_make_product = {}
    for par_var in sku_inventory_data:
        if_we_make_prod = sku_inventory_data[par_var]["if_we_make_product"]
        par_var_to_if_we_make_product[par_var] = if_we_make_prod

    for amazon_sku in amazon_sku_to_wanted_shipment_amount:
        amount_wanting_to_ship = amazon_sku_to_wanted_shipment_amount[amazon_sku]
        par_var = amazon_sku_to_par_var_dict[amazon_sku]
        cur_inv_data = sku_inventory_data[par_var]
        if_we_make_prod = cur_inv_data["if_we_make_product"]
        if if_we_make_prod == True:
            case_info = sku_inventory_data[par_var]["last_case_info"]
            quantity_per_case = case_info["quantity"]
            float_cases = float(amount_wanting_to_ship / quantity_per_case)
            int_num_of_cases = 0
            if float_cases <= 1:
                int_num_of_cases = 1
            elif is_num_whole(float_cases):
                int_num_of_cases = int(float_cases)
            else:
                low_case = m.floor(float_cases)
                high_case = m.ceil(float_cases)
                low_amount = low_case * quantity_per_case

                if low_amount >= (amount_wanting_to_ship * 0.85):
                    int_num_of_cases = low_case
                else:
                    int_num_of_cases = high_case

            inventory_to_come_up_with = int_num_of_cases * quantity_per_case
            inventory_batches = sku_inventory_data[par_var]["inventory_levels_of_different_batches"]
            if "1" in inventory_batches:
                inventory_to_come_up_with = inventory_to_come_up_with - inventory_batches["1"]["quantity"]

            if inventory_to_come_up_with > 0:
                full_sku = par_var + "-1"
                data = test_supply_levels(test_supplies_levels, supplies_needed_to_make_skus, inventory_to_come_up_with,
                                          full_sku)
                supplies_lacking = data["supplies_lacking"]
                quantity_able_to_make = data["quantity_able_to_make"]
                if quantity_able_to_make < inventory_to_come_up_with:
                    dict = {}
                    dict["supplies_lacking"] = supplies_lacking
                    amazon_skus_that_dont_have_enough_inventory[amazon_sku] = dict

                tot_inv_able_to_make = (int_num_of_cases * quantity_per_case) - (
                        inventory_to_come_up_with - quantity_able_to_make)
                int_num_of_cases = m.floor(tot_inv_able_to_make / quantity_per_case)

            # add case info to dict
            case_name = case_info["name"]
            case_weight = case_info["weight"]
            case_dimensions = case_info["dimensions"]
            item_sku = par_var + "-1"
            dict = {}
            dict["quantity_per_case"] = quantity_per_case
            dict["int_num_of_cases"] = int_num_of_cases
            dict["case_name"] = case_name
            dict["case_weight"] = case_weight
            dict["case_dimensions"] = case_dimensions
            dict["item_sku"] = item_sku

            amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku] = []
            amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku].append(dict)

        else:
            inventory_batches = sku_inventory_data[par_var]["inventory_levels_of_different_batches"]
            batch_str_nums = list(inventory_batches.keys())
            least_to_greatest_batch_nums = []
            for check_str in batch_str_nums:
                cur_num = int(check_str)
                if_added = False
                for i in range(len(least_to_greatest_batch_nums)):
                    num = least_to_greatest_batch_nums[i]
                    if cur_num < num:
                        if_added = True
                        least_to_greatest_batch_nums.insert(i, cur_num)
                        break
                if if_added == False:
                    least_to_greatest_batch_nums.append(cur_num)

            amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku] = []

            tot_inv_sent = 0

            for batch_num in least_to_greatest_batch_nums:
                batch_str = str(batch_num)
                item_sku = par_var + "-" + batch_str
                quantity_in_batch = inventory_batches[batch_str]["quantity"]
                case_info = inventory_batches[batch_str]["case_info"]

                quantity_per_case = case_info["quantity"]

                dict_to_add = {}
                dict_to_add["quantity_per_case"] = case_info["quantity"]
                dict_to_add["case_name"] = case_info["name"]
                dict_to_add["case_weight"] = case_info["weight"]
                dict_to_add["case_dimensions"] = case_info["dimensions"]
                dict_to_add["item_sku"] = item_sku

                cur_inv_needing_to_send = amount_wanting_to_ship - tot_inv_sent

                float_cases = float(cur_inv_needing_to_send / quantity_per_case)

                max_possible_amount_of_cases_to_send = m.floor(quantity_in_batch / quantity_per_case)
                if float_cases > max_possible_amount_of_cases_to_send:
                    float_cases = max_possible_amount_of_cases_to_send

                int_num_of_cases = 0
                if is_num_whole(float_cases):
                    int_num_of_cases = int(float_cases)
                else:
                    low_case = m.floor(float_cases)
                    high_case = m.ceil(float_cases)
                    low_amount = low_case * quantity_per_case

                    tot_low_amount = low_amount + tot_inv_sent

                    if tot_low_amount >= (amount_wanting_to_ship * 0.85):
                        int_num_of_cases = low_case
                    else:
                        int_num_of_cases = high_case

                dict_to_add["int_num_of_cases"] = int_num_of_cases

                amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku].append(dict_to_add)

                tot_inv_sent += (int_num_of_cases * quantity_per_case)

                if tot_inv_sent >= (amount_wanting_to_ship * 0.85):
                    break

            if tot_inv_sent < (amount_wanting_to_ship * 0.85):
                dict = {}
                dict["supplies_lacking"] = [par_var]
                amazon_skus_that_dont_have_enough_inventory[amazon_sku] = dict

    # remove cases that dont have the same quantity_per_case
    for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
        quantity_per_case_to_tot_quantity_for_that_case_type = {}
        cases = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
        for case_dict in cases:
            quantity_per_case = case_dict["quantity_per_case"]
            int_num_of_cases = case_dict["int_num_of_cases"]
            tot_q = int(quantity_per_case * int_num_of_cases)
            if quantity_per_case in quantity_per_case_to_tot_quantity_for_that_case_type:
                quantity_per_case_to_tot_quantity_for_that_case_type[quantity_per_case] = \
                    quantity_per_case_to_tot_quantity_for_that_case_type[quantity_per_case] + tot_q
            else:
                quantity_per_case_to_tot_quantity_for_that_case_type[quantity_per_case] = tot_q
        most_q = -1
        most_type = -1
        for quantity_per_case in quantity_per_case_to_tot_quantity_for_that_case_type:
            tot_q = quantity_per_case_to_tot_quantity_for_that_case_type[quantity_per_case]
            if tot_q > most_q:
                most_q = tot_q
                most_type = quantity_per_case

        list_of_types_to_remove = []
        for quantity_per_case in quantity_per_case_to_tot_quantity_for_that_case_type:
            if quantity_per_case != most_type:
                list_of_types_to_remove.append(quantity_per_case)

        for quantity_per_case in list_of_types_to_remove:
            for i in range(len(amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]) - 1, -1, -1):
                if amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku][i]["quantity_per_case"] == quantity_per_case:
                    amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku].pop(i)
    # check here that all cases for sku have same quaility
    amazon_sku_to_quantity_per_case = {}
    for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
        quantity_per_case = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku][0]["quantity_per_case"]
        amazon_sku_to_quantity_per_case[amazon_sku] = quantity_per_case

        cases = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
        for case_dict in cases:
            quantity_per_case_test = case_dict["quantity_per_case"]
            if quantity_per_case_test != quantity_per_case:
                raise Exception('quantity_per_case_test != quantity_per_case,  for amazon sku: ' + str(amazon_sku))

    # remove cases with zero quantity
    for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
        for i in range(len(amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]) - 1, -1, -1):
            if amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku][i]["int_num_of_cases"] == 0:
                amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku].pop(i)

    workbook = openpyxl.Workbook()
    view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
    workbook.views = view
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    sheet.protection.sheet = True

    orange_fill = PatternFill(start_color='FFF28C28',
                              end_color='FFF28C28',
                              fill_type='solid')

    sheet["A1"] = "not enough inventory"
    sheet['A1'].alignment = Alignment(wrap_text=True)
    sheet['A1'].fill = orange_fill
    sheet.row_dimensions[1].height = 32
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 90
    sheet.column_dimensions['D'].width = 11
    sheet.column_dimensions['E'].width = 11
    sheet.column_dimensions['F'].width = 100

    add_info_row_for_check_file(3, sheet)

    cur_row = 5
    for amazon_sku in order_of_skus_in_og_excel:
        if amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
            product_name = amazon_sku_to_title[amazon_sku]
            wanted_q = amazon_sku_to_wanted_shipment_amount[amazon_sku]
            shipping_q = 0
            for case_dict in amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]:
                quantity_per_case = case_dict["quantity_per_case"]
                int_num_of_cases = case_dict["int_num_of_cases"]
                shipping_q = shipping_q + (quantity_per_case * int_num_of_cases)

            if_highlight_row = False
            supplies_lacking = []
            if amazon_sku in amazon_skus_that_dont_have_enough_inventory:
                if_highlight_row = True
                supplies_lacking = amazon_skus_that_dont_have_enough_inventory[amazon_sku]["supplies_lacking"]

            sheet["A" + str(cur_row)] = amazon_sku_to_par_var_dict[amazon_sku]
            sheet["B" + str(cur_row)] = amazon_sku
            sheet["C" + str(cur_row)] = product_name
            sheet["D" + str(cur_row)] = wanted_q
            sheet["E" + str(cur_row)] = shipping_q
            sheet["F" + str(cur_row)] = str(supplies_lacking)
            if len(supplies_lacking) == 0:
                sheet["F" + str(cur_row)] = ""

            if if_highlight_row == True:
                for cell in sheet[str(cur_row) + ":" + str(cur_row)]:
                    cell.fill = orange_fill

            cur_row += 1

            if cur_row % 35 == 0:
                cur_row += 1
                add_info_row_for_check_file(cur_row, sheet)
                cur_row = cur_row + 2

    workbook.save(loc_of_check_excel_file)

    if_continue = True
    while True:
        print('Check the "shipment_nums_to_check.xlsx" file to make sure the numbers are good. '
              'If you want to continue with this shipment enter "yes", else enter "no"')
        x = input().lower()
        if x == "yes":
            if_continue = True
            break
        elif x == "no":
            if_continue = False
            break

    if if_continue == True:

        # remove skus with no cases
        list_skus_to_remove = []
        for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
            if len(amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]) == 0:
                list_skus_to_remove.append(amazon_sku)
        for amazon_sku in list_skus_to_remove:
            del amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]

        dict_of_amazon_sku_to_info = {}

        for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
            list_of_cases = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
            total_inventory_ship = 0
            for case in list_of_cases:
                quantity_per_case = case["quantity_per_case"]
                int_num_of_cases = case["int_num_of_cases"]
                total_inventory_ship += (quantity_per_case * int_num_of_cases)
            asin = amazon_fba_sku_dict[amazon_sku]["asin"]

            if total_inventory_ship > 0:
                dict = {}
                dict["ASIN"] = asin
                dict["Quantity"] = total_inventory_ship
                dict_of_amazon_sku_to_info[amazon_sku] = dict

        # make excel for testing cost
        workbook = openpyxl.Workbook()
        view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
        workbook.views = view
        sheet = workbook.active
        sheet.sheet_view.zoomScale = 125
        sheet.protection.sheet = True
        cur_row = 1
        for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
            total_q = 0
            list_of_cases = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
            for case in list_of_cases:
                quantity_per_case = case["quantity_per_case"]
                int_num_of_cases = case["int_num_of_cases"]
                total_q += (quantity_per_case * int_num_of_cases)
            case_dimen = list_of_cases[0]["case_dimensions"]
            case_weight = list_of_cases[0]["case_weight"]
            quantity_per_case = amazon_sku_to_quantity_per_case[amazon_sku]
            num_of_cases = total_q / quantity_per_case
            if num_of_cases % 1 != 0:
                raise Exception('num_of_cases%1 != 0')
            sheet["A" + str(cur_row)] = amazon_sku
            sheet["B" + str(cur_row)] = total_q
            sheet["E" + str(cur_row)] = quantity_per_case
            sheet["F" + str(cur_row)] = num_of_cases
            sheet["G" + str(cur_row)] = case_dimen[0]
            sheet["H" + str(cur_row)] = case_dimen[1]
            sheet["I" + str(cur_row)] = case_dimen[2]
            sheet["J" + str(cur_row)] = case_weight
            cur_row += 1
        workbook.save("excel_for_amazon_website.xlsx")
        # while True:
        #     print('Check the costs on the amazon website with the excel file, and select the best option on the api settings. Enter "done" when finished')
        #     x = input().lower()
        #     if x == "done":
        #         break

        # get location of sku's in bays
        amazon_sku_to_list_of_cases_wanting_to_ship_copy_for_func = copy.deepcopy(
            amazon_sku_to_list_of_cases_wanting_to_ship)
        get_location_of_skus(amazon_sku_to_list_of_cases_wanting_to_ship_copy_for_func)

        # make shipment here
        create_inbound_shipment(dict_of_amazon_sku_to_info, amazon_sku_to_list_of_cases_wanting_to_ship,
                                par_var_to_if_we_make_product, amazon_sku_to_quantity_per_case)

        while True:
            print('If the shipment made was different than the numbers in "shipment_nums_to_check.xlsx" '
                  'make sure to modify the supplies and sku pickle files to reflect the change after this program finishes.\nEnter "yes" to acknowledge.')
            x = input().lower()
            if x == "yes":
                break

        sku_inventory_data = pickle.load(open(loc_of_sku_inventory_file, "rb"))
        supplies_levels = pickle.load(open(loc_of_supply_inventory_file, "rb"))
        supplies_needed_to_make_skus = pickle.load(open(loc_of_supplies_needed_to_make_product, "rb"))

        for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship:
            list_of_cases = amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
            for case_dict in list_of_cases:
                item_sku = case_dict["item_sku"]
                quantity_per_case = case_dict["quantity_per_case"]
                int_num_of_cases = case_dict["int_num_of_cases"]


                index_of_second_under = item_sku.find("-", item_sku.find("-") + 1)
                par_var = item_sku[0:index_of_second_under]


                if_we_make_product = sku_inventory_data[par_var]["if_we_make_product"]
                if if_we_make_product:
                    amount_need_to_make = quantity_per_case * int_num_of_cases
                    inventory_batches = sku_inventory_data[par_var]["inventory_levels_of_different_batches"]
                    if "1" in inventory_batches:
                        amount_already_made = inventory_batches["1"]["quantity"]
                        if amount_already_made >= amount_need_to_make:
                            inventory_batches["1"]["quantity"] = amount_already_made - amount_need_to_make
                            amount_need_to_make = 0
                        else:
                            inventory_batches["1"]["quantity"] = 0
                            amount_need_to_make = amount_need_to_make - amount_already_made
                    if amount_need_to_make > 0:
                        supplies_needed_for_one = supplies_needed_to_make_skus[item_sku]
                        for supply_name in supplies_needed_for_one:
                            quant_for_supply = supplies_needed_for_one[supply_name]
                            tot = quant_for_supply * float(amount_need_to_make)
                            if supply_name not in supplies_levels:
                                print(
                                    "Error: " + supply_name + " 'supply_name not in supplies_levels' , program will make this supply go negative.")
                                supplies_levels[supply_name] = -1 * tot
                            else:
                                supplies_levels[supply_name] = supplies_levels[supply_name] - tot
                else:
                    amount_shipped = quantity_per_case * int_num_of_cases
                    inventory_batches = sku_inventory_data[par_var]["inventory_levels_of_different_batches"]

                    index_of_second_under = item_sku.find("-", item_sku.find("-") + 1)
                    batch_num = item_sku[index_of_second_under + 1:]


                    if amount_shipped > inventory_batches[batch_num]["quantity"]:
                        print(
                            "Error: batch num " + batch_num + "  of par_var: " + par_var + " went negative, program will make this batch go negative")
                    sku_inventory_data[par_var]["inventory_levels_of_different_batches"][batch_num]["quantity"] = \
                        sku_inventory_data[par_var]["inventory_levels_of_different_batches"][batch_num][
                            "quantity"] - amount_shipped

        pickle.dump(sku_inventory_data, open(loc_of_sku_inventory_file, "wb"))
        pickle.dump(supplies_levels, open(loc_of_supply_inventory_file, "wb"))
        pickle.dump(amazon_fba_sku_dict, open(loc_of_amazon_fba_sku_dict, "wb"))

        print("\n")
        print("set alarm in phone to remind about setting the pick up time with the trucking companys")
        print("\n")


def freight_class_calculator(cubic_feet, pounds):
    density = pounds / cubic_feet
    freight_class_vals = ['FC_50', 'FC_55', 'FC_60', 'FC_65', 'FC_70', 'FC_77_5', 'FC_85', 'FC_92_5', 'FC_100',
                          'FC_110', 'FC_125', 'FC_150', 'FC_175', 'FC_200', 'FC_250', 'FC_300', 'FC_400', 'FC_500']
    density_list = [
        {"high": sys.maxsize, "low": 50},
        {"high": 50, "low": 35},
        {"high": 35, "low": 30},
        {"high": 30, "low": 22.5},
        {"high": 22.5, "low": 15},
        {"high": 15, "low": 13.5},
        {"high": 13.5, "low": 12},
        {"high": 12, "low": 10.5},
        {"high": 10.5, "low": 9},
        {"high": 9, "low": 8},
        {"high": 8, "low": 7},
        {"high": 7, "low": 6},
        {"high": 6, "low": 5},
        {"high": 5, "low": 4},
        {"high": 4, "low": 3},
        {"high": 3, "low": 2},
        {"high": 2, "low": 1},
        {"high": 1, "low": 0},
    ]
    freight_class = ""
    for i in range(len(density_list)):
        high = density_list[i]["high"]
        low = density_list[i]["low"]
        if density <= high and density > low:
            freight_class = freight_class_vals[i]
            break
    if freight_class == "":
        raise Exception('freight_class == ""')

    return freight_class


def get_pallet_info_for_shipment(cases):
    inch_considered_long = 65

    total_payload_weight_pounds = 0
    total_payload_cubic_inches = 0

    total_weight_of_long_cartons = 0
    total_cubic_inches_of_long_cartons = 0
    long_case_dimensions_to_num_of_case = {}

    total_num_of_cartons = 0

    for amazon_sku in cases:
        case_list = cases[amazon_sku]
        for case_data in case_list:
            int_num_of_cases = case_data["int_num_of_cases"]
            case_dimensions = case_data["case_dimensions"]
            case_weight = case_data["case_weight"]
            case_volume = case_dimensions[0] * case_dimensions[1] * case_dimensions[2]
            total_payload_weight_pounds = total_payload_weight_pounds + (int_num_of_cases * case_weight)
            total_payload_cubic_inches = total_payload_cubic_inches + (int_num_of_cases * case_volume)

            total_num_of_cartons = total_num_of_cartons + int_num_of_cases

            str_long = ""
            if case_dimensions[0] >= inch_considered_long or case_dimensions[1] >= inch_considered_long or \
                    case_dimensions[2] >= inch_considered_long:
                list_dia = [case_dimensions[0], case_dimensions[1], case_dimensions[2]]
                list_dia.sort()
                str_long = str(list_dia[0]) + " x " + str(list_dia[1]) + " x " + str(list_dia[2])

            if str_long != "":
                total_cubic_inches_of_long_cartons = total_cubic_inches_of_long_cartons + (
                        int_num_of_cases * case_volume)
                total_weight_of_long_cartons = total_weight_of_long_cartons + (int_num_of_cases * case_weight)
                if str_long in long_case_dimensions_to_num_of_case:
                    long_case_dimensions_to_num_of_case[str_long] = long_case_dimensions_to_num_of_case[
                                                                        str_long] + 1
                else:
                    long_case_dimensions_to_num_of_case[str_long] = 1

    # get pallet list

    declared_value_per_carton = 10

    declared_value_of_shipment = declared_value_per_carton * total_num_of_cartons

    cubic_feet = total_payload_cubic_inches / (12 * 12 * 12)
    freight_class = freight_class_calculator(cubic_feet, total_payload_weight_pounds)

    standard_pallet_height = 63  # change this to try getting different prices

    height_of_wood_pallet = 5
    weight_of_wood_pallet = 15
    max_pallet_payload_weight = 1350 - weight_of_wood_pallet

    min_pallet_weight = 150

    factor_to_increase_pallet_height_by = 1.13  # this is to account for boxes not stacking perfectly

    pallet_list = []
    total_payload_inches_of_height = (total_payload_cubic_inches / (48 * 40)) * factor_to_increase_pallet_height_by

    og_total_payload_weight_pounds = total_payload_weight_pounds

    if total_cubic_inches_of_long_cartons > 0:
        # long cartons are on shipment
        print("These are the long cartons on this shipment:")
        for name in long_case_dimensions_to_num_of_case:
            print(name + " : " + str(long_case_dimensions_to_num_of_case[name]) + " cartons")

        num_of_6ft_pallets = 0
        while True:
            print("Enter how many 6ft tall pallets these cartons will make")
            while True:
                try:
                    num_of_6ft_pallets = int(input())
                    break
                except:
                    print("Enter how many 6ft tall pallets these cartons will make")
            print("enter 'yes' to confirm that these cartons will make " + str(
                num_of_6ft_pallets) + " 6ft tall pallets")
            con = input().lower()
            if con == "yes":
                break

        weight_of_each_6ft_pallet = (total_weight_of_long_cartons / num_of_6ft_pallets) + weight_of_wood_pallet

        if num_of_6ft_pallets == 1 and weight_of_each_6ft_pallet < min_pallet_weight:
            weight_of_each_6ft_pallet = min_pallet_weight

        for i in range(num_of_6ft_pallets):
            dict = {
                "weight": {
                    "unit": "LB",
                    "value": m.floor(weight_of_each_6ft_pallet)
                },
                "dimensions": {
                    "unitOfMeasurement": "IN",
                    "length": 48,
                    "width": 40,
                    "height": 72
                },
                "quantity": 1,
                "stackability": "NON_STACKABLE"
            }

            pallet_list.append(dict)

        total_payload_weight_pounds = total_payload_weight_pounds - total_weight_of_long_cartons
        total_payload_cubic_inches = total_payload_cubic_inches - total_cubic_inches_of_long_cartons
        total_payload_inches_of_height = (total_payload_cubic_inches / (48 * 40)) * factor_to_increase_pallet_height_by

    if total_payload_weight_pounds > 0:
        if total_payload_inches_of_height < (standard_pallet_height - height_of_wood_pallet):
            # make one pallet based on height
            num_of_pallets_based_on_weight = m.ceil(total_payload_weight_pounds / max_pallet_payload_weight)
            payload_height_per_pallet = total_payload_inches_of_height / num_of_pallets_based_on_weight

            height_of_pallet = payload_height_per_pallet + height_of_wood_pallet
            weight_of_pallet = (total_payload_weight_pounds / num_of_pallets_based_on_weight) + weight_of_wood_pallet
            for i in range(num_of_pallets_based_on_weight):
                dict = {
                    "weight": {
                        "unit": "LB",
                        "value": m.floor(weight_of_pallet)
                    },
                    "dimensions": {
                        "unitOfMeasurement": "IN",
                        "length": 48,
                        "width": 40,
                        "height": m.floor(height_of_pallet)
                    },
                    "quantity": 1,
                    "stackability": "NON_STACKABLE"
                }

                pallet_list.append(dict)
        else:

            num_of_pallets_based_on_height = m.ceil(
                total_payload_inches_of_height / (standard_pallet_height - height_of_wood_pallet))
            num_of_pallets_based_on_weight = m.ceil(total_payload_weight_pounds / max_pallet_payload_weight)

            num_of_pallets = num_of_pallets_based_on_height
            if num_of_pallets_based_on_weight > num_of_pallets:
                num_of_pallets = num_of_pallets_based_on_weight

            payload_height_per_pallet = total_payload_inches_of_height / num_of_pallets

            height_of_pallet = payload_height_per_pallet + height_of_wood_pallet
            weight_of_pallet = (total_payload_weight_pounds / num_of_pallets) + weight_of_wood_pallet

            for i in range(num_of_pallets):
                dict = {
                    "weight": {
                        "unit": "LB",
                        "value": m.floor(weight_of_pallet)
                    },
                    "dimensions": {
                        "unitOfMeasurement": "IN",
                        "length": 48,
                        "width": 40,
                        "height": m.floor(height_of_pallet)
                    },
                    "quantity": 1,
                    "stackability": "NON_STACKABLE"
                }
                pallet_list.append(dict)

    # check all pallet specs
    max_pallet_weight = max_pallet_payload_weight + weight_of_wood_pallet
    max_pallet_height = 72
    total_weight_pounds = 0
    number_of_pallets = len(pallet_list)

    pallet_height_to_count = {}
    for dict in pallet_list:
        height = int(dict["dimensions"]["height"])
        if height in pallet_height_to_count:
            pallet_height_to_count[height] = pallet_height_to_count[height] + 1
        else:
            pallet_height_to_count[height] = 1
        weight = int(dict["weight"]["value"])
        total_weight_pounds = total_weight_pounds + weight
        if height > max_pallet_height:
            raise Exception(
                "height: " + str(height) + " of pallet greater than max_pallet_height: " + str(max_pallet_height))
        if weight < min_pallet_weight:
            raise Exception(
                "weight: " + str(weight) + " of pallet less than min_pallet_weight: " + str(min_pallet_weight))
        if weight > max_pallet_weight:
            raise Exception(
                "weight: " + str(weight) + " of pallet more than max_pallet_weight: " + str(max_pallet_weight))

    height_ordered = list(pallet_height_to_count.keys())
    height_ordered.sort()

    text_for_pallet_info = 'Total number of pallets: ' + str(number_of_pallets) + "\n\n"

    for height in height_ordered:
        count = pallet_height_to_count[height]
        text_for_pallet_info = text_for_pallet_info + 'Number of pallets that are ' + str(height) + '" tall: ' + str(
            count) + "\n"

    return_dict = {"pallets": pallet_list,
                   "freightInformation": {"declaredValue": {"amount": str(declared_value_of_shipment), "code": "USD"},
                                          "freightClass": freight_class},
                   "text_for_pallet_info": text_for_pallet_info}
    return return_dict


def create_inbound_shipment(dict_of_amazon_sku_to_info, amazon_sku_to_list_of_cases_wanting_to_ship_og,
                            par_var_to_if_we_make_product, amazon_sku_to_quantity_per_case):

    if_time_to_finish_shipment = False
    while True:
        print("Enter 'finish' if it is time to finish the current cheapest shipping option.\nOr enter 'search' if need to search for cheapest shipment option.")
        text = input()
        if text == "finish":
            if_time_to_finish_shipment = True
            break
        elif text == "search":
            if_time_to_finish_shipment = False
            break
        else:
            r=0

    if if_time_to_finish_shipment == False:

        # delete any existing shipment info
        folder = folder_for_cheapest_shipment
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))

        # uncomment this
        date_of_pickup = ""
        while True:
            print("enter date of pick up in YYYY-MM-DD format")
            date_of_pickup = input()
            try:
                date.fromisoformat(date_of_pickup)
                break
            except ValueError:
                x = 0


        while True:


            amazon_sku_to_list_of_cases_wanting_to_ship_copy = copy.deepcopy(amazon_sku_to_list_of_cases_wanting_to_ship_og)

            inboundPlanId = ""
            inboundPlan_operationid = ""

            skus_where_none_prep_needs_to_be_assigned = {}

            while True:

                url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans"

                list_of_items = []
                for amazon_sku in dict_of_amazon_sku_to_info:
                    data = dict_of_amazon_sku_to_info[amazon_sku]
                    dict = {}
                    dict["labelOwner"] = "SELLER"
                    dict["msku"] = amazon_sku
                    dict["prepOwner"] = "SELLER"
                    if amazon_sku in skus_where_none_prep_needs_to_be_assigned:
                        dict["prepOwner"] = "NONE"
                    dict["quantity"] = data["Quantity"]
                    list_of_items.append(dict)

                payload = json.dumps({
                    "destinationMarketplaces": [market_id_USA],
                    "items": list_of_items,
                    "sourceAddress": ShipFromAddress
                })

                headers = {
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'x-amz-access-token': access_token
                }

                response = requests.request("POST", url, headers=headers, data=payload).json()

                if "errors" in response:
                    list_of_errors = response["errors"]
                    for error in list_of_errors:
                        mess = error["message"]
                        if "does not require prepOwner but SELLER was assigned. Accepted values: [NONE]" in mess:
                            sku = mess.replace("ERROR:", "").replace(
                                "does not require prepOwner but SELLER was assigned. Accepted values: [NONE]", "").strip()
                            skus_where_none_prep_needs_to_be_assigned[sku] = True
                else:
                    inboundPlan_operationid = response["operationId"]
                    inboundPlanId = response["inboundPlanId"]
                    break

            # check status of inbound request
            while True:
                try:
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + inboundPlan_operationid
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    response = requests.request("GET", url, headers=headers).json()
                    operationStatus = response["operationStatus"]
                    if operationStatus == "SUCCESS":
                        break
                    time.sleep(2)
                    print("waiting on inboundPlanId to process")
                except:
                    time.sleep(4)
                    print("in except for check status of inbound request, still working")

            # Generate packing options
            while True:
                try:
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/packingOptions"
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    response = requests.request("POST", url, headers=headers).json()
                    if "operationId" not in response:
                        raise Exception('"operationId" not in response for Generate packing options')
                    operationId = response["operationId"]
                    while True:
                        url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
                        headers = {
                            'Content-Type': 'application/json',
                            'Accept': 'application/json',
                            'x-amz-access-token': access_token
                        }

                        response = requests.request("GET", url, headers=headers).json()
                        operationStatus = response["operationStatus"]
                        if operationStatus == "SUCCESS":
                            break
                        time.sleep(2)
                        print("waiting on Generate packing options to process")
                    break
                except:
                    time.sleep(4)
                    print("in except for Generate packing options, still working")


            # list packing options
            packingOptionId = ""
            packingGroups = []
            while True:
                try:
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/packingOptions"
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    response = requests.request("GET", url, headers=headers).json()
                    if "packingOptions" in response:
                        packingGroups = response["packingOptions"][0]["packingGroups"]
                        packingOptionId = response["packingOptions"][0]["packingOptionId"]
                        break
                    time.sleep(2)
                except:
                    time.sleep(4)
                    print("in except for list packing options, still working")

            if len(packingGroups) > 1:
                # list items in each packing option
                print(
                    "here are the multiple packing groups, this program only works with one packing group, so eliminate the items from the packing groups you dont want")
                for group in packingGroups:
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/packingGroups/" + group + "/items"
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    response = requests.request("GET", url, headers=headers).json()
                    print(response)
                exit()

            packingGroupid = packingGroups[0]

            # select packing option
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/packingOptions/" + packingOptionId + "/confirmation"
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("POST", url, headers=headers).json()
            if "operationId" not in response:
                raise Exception('"operationId" not in response for select packing option')
            operationId = response["operationId"]
            while True:
                url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
                headers = {
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'x-amz-access-token': access_token
                }
                response = None
                while True:
                    response = None
                    try:
                        response = requests.request("GET", url, headers=headers).json()
                        break
                    except:
                        print("in except for waiting on select packing option to process, trying again")
                    time.sleep(3)
                operationStatus = response["operationStatus"]
                if operationStatus == "SUCCESS":
                    break
                time.sleep(2)
                print("waiting on select packing option to process")



            # set packing info
            packageGroupings = []
            packageGroupings.append({})
            packageGroupings[0]["packingGroupId"] = packingGroupid

            boxes = []
            for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship_copy:
                list_of_cartons = amazon_sku_to_list_of_cases_wanting_to_ship_copy[amazon_sku]
                for carton_data in list_of_cartons:
                    case_dimensions = carton_data["case_dimensions"]
                    case_name = carton_data["case_name"]
                    int_num_of_cases = carton_data["int_num_of_cases"]
                    case_weight = carton_data["case_weight"]
                    quantity_per_case = carton_data["quantity_per_case"]

                    data_for_boxes = {}
                    data_for_boxes["contentInformationSource"] = "BOX_CONTENT_PROVIDED"
                    # data_for_boxes["boxId"] = case_name
                    data_for_boxes["quantity"] = int_num_of_cases
                    data_for_boxes["weight"] = {"unit": "LB",
                                                "value": case_weight}
                    data_for_boxes["dimensions"] = {"height": case_dimensions[0],
                                                    "length": case_dimensions[1],
                                                    "width": case_dimensions[2],
                                                    "unitOfMeasurement": "IN"}

                    prepOwner = "SELLER"
                    if amazon_sku in skus_where_none_prep_needs_to_be_assigned:
                        prepOwner = "NONE"
                    data_for_boxes["items"] = [{"labelOwner": "SELLER",
                                                "msku": amazon_sku,
                                                "prepOwner": prepOwner,
                                                "quantity": quantity_per_case}]

                    boxes.append(data_for_boxes)
            packageGroupings[0]["boxes"] = boxes

            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/packingInformation"
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            payload = json.dumps({
                "packageGroupings": packageGroupings
            })
            response = None
            while True:
                response = None
                try:
                    response = requests.request("POST", url, headers=headers, data=payload)
                    break
                except:
                    print("in except for packing information call, trying again")
                time.sleep(3)

            status = response.status_code
            response_json = response.json()
            if status != 202 or "operationId" not in response_json:
                raise Exception('"operationId" not in response for set packing info')

            operationId = response_json["operationId"]
            while True:
                url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
                headers = {
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'x-amz-access-token': access_token
                }

                response = requests.request("GET", url, headers=headers).json()
                operationStatus = response["operationStatus"]
                if operationStatus == "SUCCESS":
                    break
                time.sleep(2)
                print("waiting on set packing info to process")



            # generate placment options
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/placementOptions"
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("POST", url, headers=headers).json()
            if "operationId" not in response:
                raise Exception('"operationId" not in response for generate placment options')

            operationId = response["operationId"]
            while True:
                url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
                headers = {
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'x-amz-access-token': access_token
                }

                response = None
                while True:
                    response = None
                    try:
                        response = requests.request("GET", url, headers=headers).json()
                        break
                    except:
                        print("in except for waiting on generate placement options to process, trying again")
                    time.sleep(3)
                operationStatus = response["operationStatus"]
                if operationStatus == "SUCCESS":
                    break
                time.sleep(2)
                print("waiting on generate placement options to process")

            # view placement option
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/placementOptions"
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("GET", url, headers=headers).json()
            placementOptions = response["placementOptions"]
            if "nextToken" in response["pagination"]:
                raise Exception('"nextToken" in response["pagination"] for view placment options')
            condensed_placement_options = []
            for data in placementOptions:
                placementOptionId = data["placementOptionId"]
                shipmentIds = data["shipmentIds"]
                fees = data["fees"]
                if len(fees) > 1:
                    raise Exception('len(fees) > 1')
                target = fees[0]["target"]
                if target != "Placement Services":
                    raise Exception('target != "Placement Services"')
                placement_fee = fees[0]["value"]["amount"]
                condensed_placement_options.append({"placementOptionId": placementOptionId,
                                                    "shipmentIds": shipmentIds,
                                                    "placement_fee": placement_fee})

            print(condensed_placement_options)
            # exit()

            # need to generate pallet info for each shipment
            # getting items per shipment
            shipment_id_to_data = {}
            for data in condensed_placement_options:
                for shipment_id in data["shipmentIds"]:

                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/shipments/" + shipment_id + "/items"
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    params = {"pageSize": "1000"}

                    response = requests.request("GET", url, headers=headers, params=params).json()
                    if "nextToken" in response["pagination"]:
                        raise Exception('"nextToken" in response["pagination"] for view items in shipment')
                    # print(json.dumps(
                    #     response,
                    #     sort_keys=True,
                    #     indent=4,
                    #     separators=(',', ': ')
                    # ))
                    items = response["items"]
                    shipment_id_to_data[shipment_id] = {}
                    shipment_id_to_data[shipment_id]["items_data"] = []
                    for item_data in items:
                        asin = item_data["asin"]
                        amazon_sku = item_data["msku"]
                        quantity = int(item_data["quantity"])
                        labelOwner = item_data["labelOwner"]
                        if labelOwner != "SELLER":
                            raise Exception('labelOwner != "SELLER"  asin: ' + asin)
                        for prep_data in item_data["prepInstructions"]:
                            prepOwner = prep_data["prepOwner"]
                            if prepOwner != "SELLER":
                                raise Exception('prepOwner != "SELLER"  asin: ' + asin)

                        quantity_per_case = amazon_sku_to_quantity_per_case[amazon_sku]
                        if quantity % quantity_per_case != 0:
                            raise Exception('quantity % quantity_per_case != 0  asin: ' + asin)

                        temp_dict = {"asin": asin,
                                     "amazon_sku": amazon_sku,
                                     "quantity": quantity}
                        shipment_id_to_data[shipment_id]["items_data"].append(temp_dict)

            # setting cases per shipment
            for data in condensed_placement_options:
                amazon_sku_to_list_of_cases_copy = copy.deepcopy(amazon_sku_to_list_of_cases_wanting_to_ship_og)
                for shipment_id in data["shipmentIds"]:
                    shipment_id_to_data[shipment_id]["cases"] = {}
                    items_data = shipment_id_to_data[shipment_id]["items_data"]
                    for item_data in items_data:
                        amazon_sku = item_data["amazon_sku"]
                        quantity = item_data["quantity"]
                        cases_to_choose_from = amazon_sku_to_list_of_cases_copy[amazon_sku]

                        shipment_id_to_data[shipment_id]["cases"][amazon_sku] = []

                        cur_quantity = quantity
                        for case_dict in cases_to_choose_from:
                            if cur_quantity < 0:
                                raise Exception('cur_quantity < 0  amazon_sku: ' + amazon_sku)
                            if cur_quantity == 0:
                                break
                            quantity_per_case = case_dict["quantity_per_case"]
                            int_num_of_cases = case_dict["int_num_of_cases"]
                            total_q_from_case = int(quantity_per_case * int_num_of_cases)
                            if total_q_from_case > cur_quantity:
                                left_over_q = total_q_from_case - cur_quantity
                                int_num_of_cases = left_over_q / quantity_per_case
                                case_dict["int_num_of_cases"] = int_num_of_cases

                                case_dict_copy = copy.deepcopy(case_dict)
                                case_dict_copy["int_num_of_cases"] = cur_quantity / quantity_per_case
                                shipment_id_to_data[shipment_id]["cases"][amazon_sku].append(case_dict_copy)
                                cur_quantity = 0
                            else:
                                case_dict["int_num_of_cases"] = 0

                                case_dict_copy = copy.deepcopy(case_dict)
                                case_dict_copy["int_num_of_cases"] = total_q_from_case / quantity_per_case
                                shipment_id_to_data[shipment_id]["cases"][amazon_sku].append(case_dict_copy)
                                cur_quantity = cur_quantity - total_q_from_case

                for amazon_sku in amazon_sku_to_list_of_cases_copy:
                    list_of_cases = amazon_sku_to_list_of_cases_copy[amazon_sku]
                    for case_dict in list_of_cases:
                        int_num_of_cases = case_dict["int_num_of_cases"]
                        if int_num_of_cases != 0:
                            raise Exception('int_num_of_cases != 0  sku: ' + amazon_sku)

            # check that the cases in the shipments equal the orginal cases needing to ship
            amazon_sku_to_list_of_cases_copy = copy.deepcopy(amazon_sku_to_list_of_cases_wanting_to_ship_og)
            amazon_sku_to_dict_of_cases_for_checking = {}
            for amazon_sku in amazon_sku_to_list_of_cases_copy:
                amazon_sku_to_dict_of_cases_for_checking[amazon_sku] = {}
                list_of_cases = amazon_sku_to_list_of_cases_copy[amazon_sku]
                for case_data in list_of_cases:
                    case_name = case_data["case_name"]
                    if case_name in amazon_sku_to_dict_of_cases_for_checking[amazon_sku]:
                        raise Exception("case_name in amazon_sku_to_dict_of_cases_for_checking[amazon_sku]")
                    amazon_sku_to_dict_of_cases_for_checking[amazon_sku][case_name] = case_data
            string_of_checking = str(json.dumps(
                amazon_sku_to_dict_of_cases_for_checking,
                sort_keys=True,
                indent=4,
                separators=(',', ': ')
            ))

            shipment_id_to_data_copy = copy.deepcopy(shipment_id_to_data)
            for data in condensed_placement_options:
                recombined_amazon_sku_to_dict_of_cases_wanting_to_ship = {}
                for shipment_id in data["shipmentIds"]:
                    cases = shipment_id_to_data_copy[shipment_id]["cases"]
                    for amazon_sku in cases:
                        case_list = copy.deepcopy(cases[amazon_sku])
                        if amazon_sku not in recombined_amazon_sku_to_dict_of_cases_wanting_to_ship:
                            recombined_amazon_sku_to_dict_of_cases_wanting_to_ship[amazon_sku] = {}
                        for case_data in case_list:
                            case_name = case_data["case_name"]
                            case_data["int_num_of_cases"] = int(case_data["int_num_of_cases"])
                            cur_num_of_cases = case_data["int_num_of_cases"]
                            if case_name in recombined_amazon_sku_to_dict_of_cases_wanting_to_ship[amazon_sku]:
                                recombined_amazon_sku_to_dict_of_cases_wanting_to_ship[amazon_sku][case_name][
                                    "int_num_of_cases"] = cur_num_of_cases + \
                                                          recombined_amazon_sku_to_dict_of_cases_wanting_to_ship[amazon_sku][
                                                              case_name]["int_num_of_cases"]
                            else:
                                recombined_amazon_sku_to_dict_of_cases_wanting_to_ship[amazon_sku][case_name] = case_data
                string_of_recombined = str(json.dumps(
                    recombined_amazon_sku_to_dict_of_cases_wanting_to_ship,
                    sort_keys=True,
                    indent=4,
                    separators=(',', ': ')
                ))
                if string_of_checking != string_of_recombined:
                    raise Exception("string_of_checking != string_of_recombined")

            # generate pallet info for each shipment
            shipment_id_to_data_copy = copy.deepcopy(shipment_id_to_data)
            for data in condensed_placement_options:
                for shipment_id in data["shipmentIds"]:
                    cases = shipment_id_to_data_copy[shipment_id]["cases"]
                    cases_copy = copy.deepcopy(cases)
                    return_dict = get_pallet_info_for_shipment(cases_copy)
                    shipment_id_to_data[shipment_id]["pallets"] = return_dict["pallets"]
                    shipment_id_to_data[shipment_id]["freightInformation"] = return_dict["freightInformation"]
                    shipment_id_to_data[shipment_id]["text_for_pallet_info"] = return_dict["text_for_pallet_info"]

            # print(json.dumps(
            #     condensed_placement_options,
            #     sort_keys=True,
            #     indent=4,
            #     separators=(',', ': ')
            # ))
            #
            # print(json.dumps(
            #     shipment_id_to_data,
            #     sort_keys=True,
            #     indent=4,
            #     separators=(',', ': ')
            # ))


            # generate transport options
            count = 0
            for placement_data in condensed_placement_options:
                # if count > 0:
                #     break
                placementOptionId = placement_data["placementOptionId"]
                shipment_ids = placement_data["shipmentIds"]

                pay_dict = {}
                pay_dict["placementOptionId"] = placementOptionId
                pay_dict["shipmentTransportationConfigurations"] = []
                for shipmentId in shipment_ids:
                    temp_dict = {}
                    temp_dict["shipmentId"] = shipmentId
                    temp_dict["readyToShipWindow"] = {"start": (date_of_pickup + "T04:00:00Z")}
                    temp_dict["contactInformation"] = {"email": Contact_details_for_shipment["Email"],
                                                       "name": Contact_details_for_shipment["Name"],
                                                       "phoneNumber": Contact_details_for_shipment["Phone"]}

                    # adding pallet info
                    # print("shipmentId: "+shipmentId+"    "+str(shipment_id_to_data[shipmentId]["pallets"]))
                    temp_dict["pallets"] = shipment_id_to_data[shipmentId]["pallets"]
                    temp_dict["freightInformation"] = shipment_id_to_data[shipmentId]["freightInformation"]

                    pay_dict["shipmentTransportationConfigurations"].append(temp_dict)

                url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/transportationOptions"
                headers = {
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'x-amz-access-token': access_token
                }
                payload = json.dumps(pay_dict)
                raw_response = requests.request("POST", url, headers=headers, data=payload)
                response = raw_response.json()
                print(response)
                if "operationId" not in response:
                    raise Exception('"operationId" not in response for generate transport options')
                operationId = response["operationId"]
                while True:
                    print("waiting on generate transport options to process for placement: " + placementOptionId)
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }

                    response = requests.request("GET", url, headers=headers).json()
                    print(response)
                    operationStatus = response["operationStatus"]
                    if operationStatus == "SUCCESS":
                        break
                    time.sleep(2)

                count += 1


            # list transport options
            shipment_id_to_options = {}

            for placement_data in condensed_placement_options:
                placementOptionId = placement_data["placementOptionId"]
                next_paginationToken = ""
                while True:
                    url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/transportationOptions"
                    headers = {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json',
                        'x-amz-access-token': access_token
                    }
                    params = {"placementOptionId": placementOptionId,
                              "pageSize": "20"}
                    if next_paginationToken != "":
                        params["paginationToken"] = next_paginationToken
                    response = requests.request("GET", url, headers=headers, params=params).json()

                    transportationOptions = response["transportationOptions"]
                    for trans_data in transportationOptions:
                        if "quote" in trans_data:
                            if trans_data["shippingMode"] == "FREIGHT_LTL":
                                price = trans_data["quote"]["cost"]["amount"]
                                shipmentId =  trans_data["shipmentId"]
                                if shipmentId not in shipment_id_to_options:
                                    shipment_id_to_options[shipmentId] = []
                                data_to_add = copy.deepcopy(trans_data)
                                if_added = False
                                for i in range(len(shipment_id_to_options[shipmentId])):
                                    data_to_check = shipment_id_to_options[shipmentId][i]
                                    price_to_check = data_to_check["quote"]["cost"]["amount"]
                                    if price < price_to_check:
                                        if_added = True
                                        shipment_id_to_options[shipmentId].insert(i,data_to_add)
                                        break
                                if if_added == False:
                                    shipment_id_to_options[shipmentId].append(data_to_add)




                    # print(json.dumps(
                    #     response,
                    #     sort_keys=True,
                    #     indent=4,
                    #     separators=(',', ': ')
                    # ))

                    if "pagination" in response:
                        if "nextToken" in response["pagination"]:
                            next_paginationToken = response["pagination"]["nextToken"]
                            if next_paginationToken == "":
                                raise Exception('next_paginationToken == ""   list transport options')
                        else:
                            break
                    else:
                        raise Exception('"pagination" not in response list transport options')




            if_select_same_carrier_for_shipment = True  # set this to false to just choose cheapest transport option for each shipment id, but the carriers will be diferent

            shipment_id_to_selected_transport_option = {}
            cheapest_plan_shipment_ids = []
            cheapest_placement_id = ""
            cheapest_ship_cost = 0
            cheapest_placment_cost = 0

            if if_select_same_carrier_for_shipment == False:

                shipment_id_to_cheapest_option = {}
                for shipment_id in shipment_id_to_options:
                    option_list = shipment_id_to_options[shipment_id]
                    for data in option_list:
                        Appointment = data["carrierAppointment"]["startTime"]
                        if date_of_pickup in Appointment:
                            data_to_add = copy.deepcopy(data)
                            shipment_id_to_cheapest_option[shipment_id] = data_to_add
                            if data_to_add["shippingSolution"] != "AMAZON_PARTNERED_CARRIER":
                                raise Exception('data_to_add["shippingSolution"] != "AMAZON_PARTNERED_CARRIER"')
                            if data_to_add["shippingMode"] != "FREIGHT_LTL":
                                raise Exception('data_to_add["shippingMode"] != "FREIGHT_LTL"')
                            break
                    price_got = shipment_id_to_cheapest_option[shipment_id]["quote"]["cost"]["amount"]
                    for data in option_list:
                        price_to_check = data["quote"]["cost"]["amount"]
                        if price_to_check < price_got:
                            raise Exception('price_to_check < price_got for shipment id: ' + shipment_id)

                # get shipment id to our cheapest transportationOptionId
                for shipment_id in shipment_id_to_cheapest_option:
                    shipment_option_data = shipment_id_to_cheapest_option[shipment_id]
                    transportationOptionId = shipment_option_data["transportationOptionId"]
                    shipment_id_to_selected_transport_option[shipment_id] = transportationOptionId

                # get cheapest placement option overall
                cheapest_total_cost = sys.float_info.max
                for placement_data in condensed_placement_options:
                    placementOptionId = placement_data["placementOptionId"]
                    shipmentIds = placement_data["shipmentIds"]
                    placement_fee = placement_data["placement_fee"]
                    total_cost = placement_fee
                    ship_cost = 0
                    for shipmentId in shipmentIds:
                        total_cost += shipment_id_to_cheapest_option[shipmentId]["quote"]["cost"]["amount"]
                        ship_cost += shipment_id_to_cheapest_option[shipmentId]["quote"]["cost"]["amount"]
                    if total_cost < cheapest_total_cost:
                        cheapest_total_cost = total_cost
                        cheapest_placement_id = placementOptionId
                        cheapest_ship_cost = ship_cost
                        cheapest_placment_cost = placement_fee
                        cheapest_plan_shipment_ids = copy.deepcopy(shipmentIds)
            else:

                shipment_id_to_dict_of_carriers = {}
                for shipment_id in shipment_id_to_options:
                    shipment_id_to_dict_of_carriers[shipment_id] = {}
                    option_list = shipment_id_to_options[shipment_id]
                    for data in option_list:
                        Appointment = data["carrierAppointment"]["startTime"]
                        if date_of_pickup in Appointment:
                            data_to_add = copy.deepcopy(data)
                            if data_to_add["shippingSolution"] != "AMAZON_PARTNERED_CARRIER":
                                raise Exception('data_to_add["shippingSolution"] != "AMAZON_PARTNERED_CARRIER"')
                            if data_to_add["shippingMode"] != "FREIGHT_LTL":
                                raise Exception('data_to_add["shippingMode"] != "FREIGHT_LTL"')

                            alphaCode = data_to_add["carrier"]["alphaCode"]
                            if alphaCode in shipment_id_to_dict_of_carriers[shipment_id]:
                                raise Exception('duplicate appointments at same time from same carrier')
                            shipment_id_to_dict_of_carriers[shipment_id][alphaCode] = data_to_add

                    for data in option_list:
                        price_to_check = data["quote"]["cost"]["amount"]
                        alphaCode = data["carrier"]["alphaCode"]

                        lowest_price_got_for_this_carrier = \
                            shipment_id_to_dict_of_carriers[shipment_id][alphaCode]["quote"]["cost"]["amount"]
                        if price_to_check < lowest_price_got_for_this_carrier:
                            raise Exception(
                                'price_to_check < lowest_price_got_for_this_carrier for shipment id: ' + shipment_id)

                # print(json.dumps(
                #     shipment_id_to_dict_of_carriers,
                #     sort_keys=True,
                #     indent=4,
                #     separators=(',', ': ')
                # ))

                # get cheapest placement option overall
                cheapest_total_cost = sys.float_info.max
                chosen_carrier = ""
                for placement_data in condensed_placement_options:
                    placementOptionId = placement_data["placementOptionId"]
                    shipmentIds = placement_data["shipmentIds"]
                    placement_fee = placement_data["placement_fee"]
                    total_cost = placement_fee
                    ship_cost = 0

                    dict_of_carrier_to_tot_cost_and_shipments_apart_of = {}

                    for shipmentId in shipmentIds:
                        dict_of_carriers = shipment_id_to_dict_of_carriers[shipmentId]
                        for carrier in dict_of_carriers:
                            if carrier not in dict_of_carrier_to_tot_cost_and_shipments_apart_of:
                                dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier] = {}
                                dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier]["tot_cost"] = 0
                                dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier]["shipments_carrier_can_do"] = {}
                            cost_for_this_carrier_shipment = dict_of_carriers[carrier]["quote"]["cost"]["amount"]
                            dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier]["shipments_carrier_can_do"][
                                shipmentId] = True
                            dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier][
                                "tot_cost"] += cost_for_this_carrier_shipment

                    lowest_cost_carrier = ""
                    lowest_cost = sys.float_info.max
                    for carrier in dict_of_carrier_to_tot_cost_and_shipments_apart_of:
                        if_carrier_can_do_all_shipments = True
                        for shipmentId in shipmentIds:
                            if shipmentId not in dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier][
                                "shipments_carrier_can_do"]:
                                if_carrier_can_do_all_shipments = False
                        if if_carrier_can_do_all_shipments == True:
                            cur_cost = dict_of_carrier_to_tot_cost_and_shipments_apart_of[carrier]["tot_cost"]
                            if cur_cost < lowest_cost:
                                lowest_cost = cur_cost
                                lowest_cost_carrier = carrier

                    ship_cost = lowest_cost
                    total_cost += ship_cost
                    if total_cost < cheapest_total_cost:
                        cheapest_total_cost = total_cost
                        cheapest_placement_id = placementOptionId
                        cheapest_ship_cost = ship_cost
                        cheapest_placment_cost = placement_fee
                        cheapest_plan_shipment_ids = copy.deepcopy(shipmentIds)
                        chosen_carrier = lowest_cost_carrier

                # check total ship cost
                total_ship_check = 0
                for shipment_id in cheapest_plan_shipment_ids:
                    total_ship_check += shipment_id_to_dict_of_carriers[shipment_id][chosen_carrier]["quote"]["cost"]["amount"]
                if total_ship_check != cheapest_ship_cost:
                    raise Exception('total_ship_check != cheapest_ship_cost')

                # get shipment id to our selected transportationOptionId
                for shipment_id in cheapest_plan_shipment_ids:
                    shipment_option_data = shipment_id_to_dict_of_carriers[shipment_id][chosen_carrier]
                    transportationOptionId = shipment_option_data["transportationOptionId"]
                    shipment_id_to_selected_transport_option[shipment_id] = transportationOptionId

            total_payload_weight = 0
            for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship_copy:
                list_of_cartons = amazon_sku_to_list_of_cases_wanting_to_ship_copy[amazon_sku]
                for carton_data in list_of_cartons:
                    int_num_of_cases = carton_data["int_num_of_cases"]
                    case_weight = carton_data["case_weight"]
                    total_payload_weight = total_payload_weight + (int_num_of_cases * case_weight)

            print("cheapest placement id: " + str(cheapest_placement_id))
            print("total weight of payload: " + str(total_payload_weight))
            print("cheapest ship cost: $" + str(cheapest_ship_cost))
            print("cheapest placment cost: $" + str(cheapest_placment_cost))




            # record_info_in_pickle

            print("writing started")
            # temp file setup
            cur_data = []
            if os.path.exists(path_to_data_file_for_cheapest_shipment):
                cur_data = pickle.load(open(path_to_data_file_for_cheapest_shipment, "rb"))
            new_dict = {
                "cheapest_plan_shipment_ids": cheapest_plan_shipment_ids,
                "shipment_id_to_data": shipment_id_to_data,
                "inboundPlanId": inboundPlanId,
                "amazon_sku_to_list_of_cases_wanting_to_ship_og": amazon_sku_to_list_of_cases_wanting_to_ship_og,
                "shipment_id_to_selected_transport_option": shipment_id_to_selected_transport_option,
                "cheapest_placement_id": cheapest_placement_id,
                "total_payload_weight": total_payload_weight,
                "cheapest_ship_cost": cheapest_ship_cost,
                "cheapest_placment_cost": cheapest_placment_cost
            }
            cur_data.append(new_dict)
            while True:
                try:
                    pickle.dump(cur_data, open(path_to_data_file_for_cheapest_shipment, "wb"))

                    test_read = pickle.load(open(path_to_data_file_for_cheapest_shipment, "rb"))
                    break
                except Exception as e:
                    print(e)
                    print("issue writing cur_data file, trying again")

            cheapest_overall_tot_cost = sys.float_info.max
            cheapest_overall_ship_cost = 0
            cheapest_overall_placement_cost = 0
            total_payload_weight_testing = 0
            for data in cur_data:
                cheapest_ship_cost = data["cheapest_ship_cost"]
                cheapest_placment_cost = data["cheapest_placment_cost"]
                total_payload_weight = data["total_payload_weight"]
                if total_payload_weight_testing == 0:
                    total_payload_weight_testing = total_payload_weight
                if total_payload_weight != total_payload_weight_testing:
                    raise Exception('total_payload_weight != total_payload_weight_testing')
                test_tot_cost = cheapest_ship_cost + cheapest_placment_cost
                if test_tot_cost < cheapest_overall_tot_cost:
                    cheapest_overall_tot_cost = test_tot_cost
                    cheapest_overall_ship_cost = cheapest_ship_cost
                    cheapest_overall_placement_cost = cheapest_placment_cost


            # write text file
            f = open(path_to_file_for_cheapest_shipment_info_text, "w")
            f.write("payload_weight: " + str(total_payload_weight_testing))
            f.write("\ncheapest_overall_tot_cost: $" + str(cheapest_overall_tot_cost))
            f.write("\ncheapest_overall_ship_cost: $" + str(cheapest_overall_ship_cost))
            f.write("\ncheapest_overall_placement_cost: $" + str(cheapest_overall_placement_cost))
            f.close()

            print("writing done")
            print("\n\n")

            print("payload_weight: "+str(total_payload_weight_testing))
            print("cheapest_overall_tot_cost: $"+str(cheapest_overall_tot_cost))
            print("cheapest_overall_ship_cost: $"+str(cheapest_overall_ship_cost))
            print("cheapest_overall_placement_cost: $"+str(cheapest_overall_placement_cost))








    else:
        # finishing shipment

        # read file information
        cur_data = pickle.load(open(path_to_data_file_for_cheapest_shipment, "rb"))


        cheapest_tot_cost = sys.float_info.max
        cheapest_data = {}
        for data in cur_data:
            cheapest_ship_cost = data["cheapest_ship_cost"]
            cheapest_placment_cost = data["cheapest_placment_cost"]
            tot = cheapest_ship_cost+cheapest_placment_cost
            if tot < cheapest_tot_cost:
                cheapest_tot_cost = tot
                cheapest_data = data

        cheapest_plan_shipment_ids = cheapest_data["cheapest_plan_shipment_ids"]
        shipment_id_to_data = cheapest_data["shipment_id_to_data"]
        inboundPlanId = cheapest_data["inboundPlanId"]
        amazon_sku_to_list_of_cases_wanting_to_ship_og = cheapest_data["amazon_sku_to_list_of_cases_wanting_to_ship_og"]
        shipment_id_to_selected_transport_option = cheapest_data["shipment_id_to_selected_transport_option"]
        cheapest_placement_id = cheapest_data["cheapest_placement_id"]
        total_payload_weight = cheapest_data["total_payload_weight"]
        cheapest_ship_cost = cheapest_data["cheapest_ship_cost"]
        cheapest_placment_cost = cheapest_data["cheapest_placment_cost"]

        calc_tot_cost = cheapest_ship_cost+cheapest_placment_cost
        if calc_tot_cost != cheapest_tot_cost:
            raise Exception('calc_tot_cost != cheapest_tot_cost')


        print("payload_weight: " + str(total_payload_weight))
        print("cheapest_overall_tot_cost: $" + str(calc_tot_cost))
        print("cheapest_overall_ship_cost: $" + str(cheapest_ship_cost))
        print("cheapest_overall_placement_cost: $" + str(cheapest_placment_cost))



        print("")
        while True:
            print('Enter "next" if you want to proceed with shipment')
            x = input().lower()
            if x == "next":
                break



        # # confirm placment option
        url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/placementOptions/" + cheapest_placement_id + "/confirmation"
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'x-amz-access-token': access_token
        }
        raw_response = requests.request("POST", url, headers=headers)
        response = raw_response.json()
        # print(response)
        if "operationId" not in response:
            raise Exception('"operationId" not in response for confirm placment option')
        operationId = response["operationId"]
        while True:
            print("waiting on confirm placment option to process")
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }

            response = requests.request("GET", url, headers=headers).json()
            print(response)
            operationStatus = response["operationStatus"]
            if operationStatus == "SUCCESS":
                break
            time.sleep(2)



        # select/pay for transport options
        url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/transportationOptions/confirmation"
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'x-amz-access-token': access_token
        }
        list_of_selections = []
        for shipment_id in cheapest_plan_shipment_ids:
            dict_to_add = {}
            dict_to_add["shipmentId"] = shipment_id
            dict_to_add["transportationOptionId"] = shipment_id_to_selected_transport_option[shipment_id]
            dict_to_add["contactInformation"] = {"email": Contact_details_for_shipment["Email"],
                                                 "name": Contact_details_for_shipment["Name"],
                                                 "phoneNumber": Contact_details_for_shipment["Phone"]}
            list_of_selections.append(dict_to_add)

        pay_dict = {"transportationSelections": list_of_selections}
        print(json.dumps(
            pay_dict,
            sort_keys=True,
            indent=4,
            separators=(',', ': ')
        ))
        payload = json.dumps(pay_dict)
        print(pay_dict)
        raw_response = requests.request("POST", url, headers=headers, data=payload)
        response = raw_response.json()
        print(response)
        if "operationId" not in response:
            raise Exception('"operationId" not in response for select/pay for transport options')
        operationId = response["operationId"]
        while True:
            print("waiting on select/pay for transport options to process")
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }

            response = requests.request("GET", url, headers=headers).json()
            print(json.dumps(
                response,
                sort_keys=True,
                indent=4,
                separators=(',', ': ')
            ))
            operationStatus = response["operationStatus"]
            if operationStatus == "SUCCESS":
                break
            time.sleep(2)



        # cancel shipment charges
        # url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/"+inboundPlanId+"/cancellation"
        # headers = {
        #     'Content-Type': 'application/json',
        #     'Accept': 'application/json',
        #     'x-amz-access-token': access_token
        # }
        # raw_response = requests.request("POST", url, headers=headers)
        # response = raw_response.json()
        # print(response)
        # if "operationId" not in response:
        #     raise Exception('"operationId" not in response for cancelling placement option')
        # operationId = response["operationId"]
        # while True:
        #     print("waiting on cancelling placement option to process")
        #     url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/operations/" + operationId
        #     headers = {
        #         'Content-Type': 'application/json',
        #         'Accept': 'application/json',
        #         'x-amz-access-token': access_token
        #     }
        #
        #     response = requests.request("GET", url, headers=headers).json()
        #     print(response)
        #     operationStatus = response["operationStatus"]
        #     if operationStatus == "SUCCESS":
        #         break
        #     time.sleep(2)
        # print("placement option "+inboundPlanId+" succuesfully cancelled")
        # exit()




        # get pallet and package labels

        # *
        # *
        # *
        # *
        # *
        # WARNING WARNING WARNING,  case data in "shipment_id_to_data" may not be the exact diamensions of cases that is return by calling the
        # ship "/boxes" call.   We will assign C case numbers to boxes based off the "/boxes" call
        # *
        # *
        # *
        # *
        # *

        shipment_id_to_old_way_total_carton_num = {}
        for shipment_id in shipment_id_to_data:
            case_dict = shipment_id_to_data[shipment_id]["cases"]
            number_of_cartons = 0
            for amazon_sku in case_dict:
                case_list = case_dict[amazon_sku]
                for case_data in case_list:
                    int_num_of_cases = case_data["int_num_of_cases"]
                    number_of_cartons += int_num_of_cases
            shipment_id_to_old_way_total_carton_num[shipment_id] = number_of_cartons
            del shipment_id_to_data[shipment_id]["cases"]

        name_of_folder_for_info = str(datetime.now()).replace(" ", "_").replace(".", "_").replace(":", "-")
        os.mkdir(loc_of_shipment_info + "/" + name_of_folder_for_info)
        root_path_to_info = loc_of_shipment_info + "/" + name_of_folder_for_info

        amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from = copy.deepcopy(
            amazon_sku_to_list_of_cases_wanting_to_ship_og)

        # print(json.dumps(
        #     amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from,
        #     sort_keys=True,
        #     indent=4,
        #     separators=(',', ': ')
        # ))

        plan_shipment_id_to_actual_shipment_id = {}
        for plan_shipment_id in cheapest_plan_shipment_ids:

            # get shipment details
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/shipments/" + plan_shipment_id
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("GET", url, headers=headers).json()
            shipmentConfirmationId = response["shipmentConfirmationId"]
            plan_shipment_id_to_actual_shipment_id[plan_shipment_id] = shipmentConfirmationId

            os.mkdir(root_path_to_info + "/" + shipmentConfirmationId)
            path_to_shipment = root_path_to_info + "/" + shipmentConfirmationId

            # pallet labels
            shipment_data = shipment_id_to_data[plan_shipment_id]
            pallet_list = shipment_data["pallets"]
            num_of_pallets = 0
            for pallet_data in pallet_list:
                num_of_pallets += pallet_data["quantity"]

            url = "https://sellingpartnerapi-na.amazon.com/fba/inbound/v0/shipments/" + shipmentConfirmationId + "/labels"

            params = {"PageType": "PackageLabel_Letter_6",
                      "LabelType": "PALLET",
                      "NumberOfPallets": str(num_of_pallets)}

            headers = {
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }

            response = requests.request("GET", url, headers=headers, params=params)
            code = str(response.status_code)
            download_url = str(response.json()["payload"]["DownloadURL"])
            if code != "200":
                raise Exception('code != "200" for pallet labels')

            urllib.request.urlretrieve(download_url, path_to_shipment + "/pallet_labels_" + shipmentConfirmationId + ".pdf")

            # box labels
            # listShipmentBoxes
            url = "https://sellingpartnerapi-na.amazon.com/inbound/fba/2024-03-20/inboundPlans/" + inboundPlanId + "/shipments/" + plan_shipment_id + "/boxes"
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }
            response = requests.request("GET", url, headers=headers).json()
            if "nextToken" in response["pagination"]:
                raise Exception('"nextToken" in response["pagination"] for /boxes call')

            # print(json.dumps(
            #     response,
            #     sort_keys=True,
            #     indent=4,
            #     separators=(',', ': ')
            # ))

            case_name_to_amazon_sku = {}

            items_data_for_shipment_to_delete = copy.deepcopy(shipment_data["items_data"])

            # map amazon boxid to our case name. And remove from amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from
            boxid_to_case_name = {}
            boxes = response["boxes"]
            for box_data in boxes:
                boxId = box_data["boxId"]
                items = box_data["items"]
                if len(items) != 1:
                    raise Exception('len(items) != 1')

                prepInstructions = items[0]["prepInstructions"]
                for prep in prepInstructions:
                    prepOwner = prep["prepOwner"]
                    if prepOwner != "SELLER":
                        raise Exception('prepOwner != "SELLER"')

                if items[0]["labelOwner"] != "SELLER":
                    raise Exception('items[0]["labelOwner"] != "SELLER"')

                if box_data["quantity"] != 1:
                    raise Exception('box_data["quantity"] != 1')

                dimensions = box_data["dimensions"]
                weight = box_data["weight"]

                if dimensions["unitOfMeasurement"] != "IN":
                    raise Exception('dimensions["unitOfMeasurement"] != "IN"')

                if weight["unit"] != "LB":
                    raise Exception('weight["unit"] != "LB"')

                amazon_sku = items[0]["msku"]
                quantity_of_items_in_case = items[0]["quantity"]
                case_weight_pounds = weight["value"]
                case_dimensions_list_inches = [dimensions["height"], dimensions["length"], dimensions["width"]]
                case_dimensions_list_inches.sort()

                # delete from items_data_for_shipment_to_delete
                for item_dict in items_data_for_shipment_to_delete:
                    our_item_amazon_sku = item_dict["amazon_sku"]
                    if amazon_sku == our_item_amazon_sku:
                        item_dict["quantity"] = item_dict["quantity"] - quantity_of_items_in_case
                        break

                if_found_our_case = False

                list_of_our_cases_to_choose_from = amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from[amazon_sku]
                spot_in_case_list = 0
                if_need_to_delete_case_from_list = False
                for our_case_data in list_of_our_cases_to_choose_from:
                    our_case_dimensions = our_case_data["case_dimensions"]
                    our_case_weight = our_case_data["case_weight"]
                    our_quantity_per_case = our_case_data["quantity_per_case"]
                    our_case_name = our_case_data["case_name"]
                    our_case_dimensions.sort()
                    if case_dimensions_list_inches[0] == our_case_dimensions[0] and case_dimensions_list_inches[1] == \
                            our_case_dimensions[1] and case_dimensions_list_inches[2] == our_case_dimensions[2]:
                        if case_weight_pounds == our_case_weight:
                            if quantity_of_items_in_case == our_quantity_per_case:
                                if_found_our_case = True
                                if boxId in boxid_to_case_name:
                                    raise Exception('boxId in boxid_to_case_name')
                                boxid_to_case_name[boxId] = our_case_name
                                int_num_of_cases = our_case_data["int_num_of_cases"] - 1
                                our_case_data["int_num_of_cases"] = int_num_of_cases
                                if int_num_of_cases < 0:
                                    raise Exception('int_num_of_cases < 0')

                                if int_num_of_cases == 0:
                                    if_need_to_delete_case_from_list = True

                                if our_case_name in case_name_to_amazon_sku:
                                    check_amazon_sku = case_name_to_amazon_sku[our_case_name]
                                    if check_amazon_sku != amazon_sku:
                                        raise Exception('check_amazon_sku != amazon_sku')
                                case_name_to_amazon_sku[our_case_name] = amazon_sku

                                break

                    spot_in_case_list += 1

                if if_found_our_case == False:
                    raise Exception('if if_found_our_case == False')

                if if_need_to_delete_case_from_list == True:
                    amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from[amazon_sku].pop(spot_in_case_list)

            for item_dict in items_data_for_shipment_to_delete:
                quantity_to_test = item_dict["quantity"]
                if quantity_to_test != 0:
                    raise Exception('quantity_to_test != 0')

            box_ids_to_print_labels_for = list(boxid_to_case_name.keys())
            total_num_of_cartons_from_box_call = len(box_ids_to_print_labels_for)
            total_num_of_cartons_old_way = shipment_id_to_old_way_total_carton_num[plan_shipment_id]
            if total_num_of_cartons_from_box_call != total_num_of_cartons_old_way:
                raise Exception('total_num_of_cartons_from_box_call != total_num_of_cartons_old_way')

            # get box labels

            url = "https://sellingpartnerapi-na.amazon.com/fba/inbound/v0/shipments/" + shipmentConfirmationId + "/labels"

            params = {"PageType": "PackageLabel_Letter_6",
                      "LabelType": "UNIQUE",
                      "PackageLabelsToPrint": ','.join(box_ids_to_print_labels_for),
                      "NumberOfPackages": total_num_of_cartons_from_box_call}

            headers = {
                'Accept': 'application/json',
                'x-amz-access-token': access_token
            }

            response = requests.request("GET", url, headers=headers, params=params)
            code = str(response.status_code)
            download_url = str(response.json()["payload"]["DownloadURL"])
            if code != "200":
                raise Exception('code != "200" for package labels')
            urllib.request.urlretrieve(download_url, path_to_shipment + "/package_labels.pdf")

            padded_carton_num_to_case_name = {}
            for full_box_id in boxid_to_case_name:
                case_name = boxid_to_case_name[full_box_id]
                padded_carton_num = full_box_id.replace(shipmentConfirmationId, "")
                if len(padded_carton_num) != 7:
                    raise Exception('len(padded_carton_num) != 7')
                if padded_carton_num[0] != "U":
                    raise Exception('padded_carton_num[0] != "U"')
                if padded_carton_num in padded_carton_num_to_case_name:
                    raise Exception('padded_carton_num in padded_carton_num_to_case_name')
                padded_carton_num_to_case_name[padded_carton_num] = case_name

            package_label_path = path_to_shipment + "/package_labels.pdf"
            package_label_with_case_name_path = path_to_shipment + "/carton_labels_with_case_names_" + shipmentConfirmationId + ".pdf"
            add_case_names_to_carton_labels(padded_carton_num_to_case_name, shipmentConfirmationId, package_label_path,
                                            package_label_with_case_name_path, case_name_to_amazon_sku)
            os.remove(package_label_path)

            # get item labels

            amazon_sku_to_list_of_cases_wanting_to_ship_another_copy = copy.deepcopy(
                amazon_sku_to_list_of_cases_wanting_to_ship_og)

            cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship = {}
            correct_case_name_to_count = {}
            for box_id in boxid_to_case_name:
                cur_case_name = boxid_to_case_name[box_id]
                if cur_case_name not in correct_case_name_to_count:
                    correct_case_name_to_count[cur_case_name] = 0
                correct_case_name_to_count[cur_case_name] += 1
                for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship_another_copy:
                    list_of_cases = amazon_sku_to_list_of_cases_wanting_to_ship_another_copy[amazon_sku]
                    for case_data in list_of_cases:
                        case_name = case_data["case_name"]
                        if cur_case_name == case_name:
                            if amazon_sku not in cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship:
                                cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku] = []
                            # check if case already exist
                            if_case_already_exists = False
                            checking_list_of_case = cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
                            for check_case_data in checking_list_of_case:
                                checking_case_name = check_case_data["case_name"]
                                if checking_case_name == cur_case_name:
                                    if_case_already_exists = True
                                    check_case_data["int_num_of_cases"] = check_case_data["int_num_of_cases"] + 1
                                    break
                            if if_case_already_exists == False:
                                case_data_copy = copy.deepcopy(case_data)
                                case_data_copy["int_num_of_cases"] = 1
                                cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku].append(case_data_copy)

            # checking that cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship is correct
            items_data_for_shipment_copy = copy.deepcopy(shipment_data["items_data"])

            added_up_amazon_sku_to_quanity = {}
            for amazon_sku in cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship:
                total_q_for_sku = 0
                case_list = cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
                for case_data in case_list:
                    case_name = case_data["case_name"]
                    count_of_cases = case_data["int_num_of_cases"]
                    q_per_case = case_data["quantity_per_case"]
                    total_q_for_sku = total_q_for_sku + (count_of_cases * q_per_case)
                    if correct_case_name_to_count[case_name] != count_of_cases:
                        raise Exception(
                            'if correct_case_name_to_count[case_name] != count_of_cases for amazon sku: ' + amazon_sku)
                added_up_amazon_sku_to_quanity[amazon_sku] = total_q_for_sku
            for item_data in items_data_for_shipment_copy:
                amazon_sku = item_data["amazon_sku"]
                quantity_shipping = item_data["quantity"]
                if added_up_amazon_sku_to_quanity[amazon_sku] != quantity_shipping:
                    raise Exception(
                        'if added_up_amazon_sku_to_quanity[amazon_sku] != quantity_shipping for amazon sku: ' + amazon_sku)

            sku_list = []
            sku_dict = {}
            for item_data in items_data_for_shipment_copy:
                amazon_sku = item_data["amazon_sku"]
                sku_list.append(amazon_sku)
                if amazon_sku in sku_dict:
                    raise Exception('if amazon_sku in sku_dict for amazon sku: ' + amazon_sku)
                sku_dict[amazon_sku] = True
            list_catalog_data = get_inventory(sku_list)

            amazon_sku_to_needed_data = {}
            for cur_data in list_catalog_data:
                amazon_sku = cur_data["sellerSku"]
                condition = cur_data["condition"]
                if condition == "NewItem":
                    condition = "New"
                else:
                    raise Exception('condition != "NewItem" for amazon sku: ' + amazon_sku)
                title = cur_data["productName"]
                fnsku = cur_data["fnSku"]
                amazon_sku_to_needed_data[amazon_sku] = {"condition": condition, "title": title, "fnsku": fnsku}

            our_sku_list = []
            fnsku_list = []
            title_list = []
            condition_list = []
            num_of_products_shipping_list = []
            list_of_items_per_carton = []

            carton_name_to_info_for_excel_sheet = {}

            for item_data in items_data_for_shipment_copy:
                amazon_sku = item_data["amazon_sku"]
                tot_quantity_shipping = item_data["quantity"]

                tot_quantity_according_to_cases = 0
                cases_wanting_to_ship = cur_shipment_amazon_sku_to_list_of_cases_wanting_to_ship[amazon_sku]
                for case_info in cases_wanting_to_ship:
                    case_name = case_info["case_name"]
                    our_sku = case_info["item_sku"]
                    quantity_per_case = case_info["quantity_per_case"]
                    int_num_of_cases = case_info["int_num_of_cases"]
                    case_weight_lbs = case_info["case_weight"]
                    case_dimensions_inch = case_info["case_dimensions"]
                    cur_quantity = quantity_per_case * int_num_of_cases
                    tot_quantity_according_to_cases += cur_quantity


                    index_of_second_under = our_sku.find("-", our_sku.find("-") + 1)
                    cur_par_var = our_sku[0:index_of_second_under]

                    if_we_make_product = par_var_to_if_we_make_product[cur_par_var]

                    amazon_data = amazon_sku_to_needed_data[amazon_sku]
                    fnsku = amazon_data["fnsku"]
                    condition = amazon_data["condition"]
                    title = amazon_data["title"]

                    our_sku_list.append(our_sku)
                    fnsku_list.append(fnsku)
                    title_list.append(title)
                    condition_list.append(condition)
                    num_of_products_shipping_list.append(cur_quantity)
                    list_of_items_per_carton.append(quantity_per_case)

                    carton_name_to_info_for_excel_sheet[case_name] = {"int_num_of_cases": int_num_of_cases,
                                                                      "product_quantity": cur_quantity,
                                                                      "case_weight_lbs": case_weight_lbs,
                                                                      "case_dimensions_inch": case_dimensions_inch,
                                                                      "if_we_make_product": if_we_make_product,
                                                                      "title": title,
                                                                      "our_sku": our_sku
                                                                      }

                if tot_quantity_according_to_cases != tot_quantity_shipping:
                    raise Exception(
                        'tot_quantity_according_to_cases != tot_quantity_shipping for amazon sku: ' + amazon_sku)

            file_path_name = path_to_shipment + "/product_labels_" + shipmentConfirmationId + ".pdf"
            product_label_pdf(our_sku_list, fnsku_list, title_list, condition_list, num_of_products_shipping_list,
                              file_path_name, shipmentConfirmationId, list_of_items_per_carton)

            # make shipment summary excel
            file_path_for_excel_summary = path_to_shipment + "/summary_" + shipmentConfirmationId + ".xlsx"
            make_excel_summary_file_for_shipment(carton_name_to_info_for_excel_sheet, shipmentConfirmationId,
                                                 file_path_for_excel_summary)

            # make pallet info text file
            text_for_pallet_info = shipment_data["text_for_pallet_info"]
            with open(path_to_shipment + '/pallet_info_' + shipmentConfirmationId + '.txt', 'w') as f:
                f.write(text_for_pallet_info)
                f.close()

        # checking that all our cases we wanted to ship got labels
        for amazon_sku in amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from:
            list_of_cases = amazon_sku_to_list_of_cases_wanting_to_ship_to_delete_from[amazon_sku]
            if len(list_of_cases) != 0:
                raise Exception('len(list_of_cases) != 0')

        # make total summary excel
        try:
            make_total_summary_excel(name_of_folder_for_info)
        except Exception as e:
            print(e)
            print("ERROR making total summary excel file")

        print("\n")

        print(
            "Print all label pdf files on a windows operations system and use the chrome browser to print from.\nMake sure the scale in print settings is set to 100%\n")





def labor_for_sku_labeling_calculator(item_weight_lbs, item_volume_inch):
    dimensional_weight = item_volume_inch / 139
    weight_to_use = item_weight_lbs
    if dimensional_weight > weight_to_use:
        weight_to_use = dimensional_weight
    time_per_item_in_hours = 0.002666 + (weight_to_use * 0.000916)
    cost_per_item = time_per_item_in_hours * 30
    return cost_per_item


def make_excel_summary_file_for_shipment(carton_name_to_info_for_excel_sheet, shipment_id, file_path):
    workbook = openpyxl.Workbook()
    view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
    workbook.views = view
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    sheet.protection.sheet = True

    sheet["A1"] = "SKU"
    sheet["B1"] = "Carton Name"
    sheet["C1"] = "Carton Quantity"
    sheet["D1"] = "Products Per Carton"
    sheet["E1"] = "Product Quantity"
    sheet["F1"] = "SKU Labor Cost Per Item"
    sheet["G1"] = "Location in Bays"
    sheet["H1"] = "Title"
    sheet['A1'].alignment = Alignment(wrap_text=True)
    sheet['B1'].alignment = Alignment(wrap_text=True)
    sheet['C1'].alignment = Alignment(wrap_text=True)
    sheet['D1'].alignment = Alignment(wrap_text=True)
    sheet['E1'].alignment = Alignment(wrap_text=True)
    sheet['F1'].alignment = Alignment(wrap_text=True)
    sheet['G1'].alignment = Alignment(wrap_text=True)
    sheet['H1'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[1].height = 32
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12

    total_sku_labeling_labor = 0

    cur_row = 2
    for carton_name in carton_name_to_info_for_excel_sheet:
        data = carton_name_to_info_for_excel_sheet[carton_name]
        int_num_of_cases = data["int_num_of_cases"]
        product_quantity = data["product_quantity"]
        products_per_carton = product_quantity / int_num_of_cases
        our_sku = data["our_sku"]
        title = data["title"]

        location_in_bays = ""
        sku_for_bay = carton_name.replace("C", "").strip()
        if sku_for_bay in sku_to_location_in_bays:
            location_in_bays = sku_to_location_in_bays[sku_for_bay]

        if_we_make_product = data["if_we_make_product"]
        if if_we_make_product == False:
            item_weight_lbs = data["case_weight_lbs"] / products_per_carton
            case_dimensions_inch = data["case_dimensions_inch"]
            case_volume = case_dimensions_inch[0] * case_dimensions_inch[1] * case_dimensions_inch[2]
            item_volume_inch = case_volume / products_per_carton
            cost_per_item = labor_for_sku_labeling_calculator(item_weight_lbs, item_volume_inch)
            total_sku_labeling_labor += cost_per_item * product_quantity
            sheet["F" + str(cur_row)] = round(cost_per_item, 4)

        sheet["A" + str(cur_row)] = "SKU " + our_sku
        sheet["B" + str(cur_row)] = carton_name
        sheet["C" + str(cur_row)] = int_num_of_cases
        sheet["D" + str(cur_row)] = products_per_carton
        sheet["E" + str(cur_row)] = product_quantity
        sheet["G" + str(cur_row)] = location_in_bays
        sheet["H" + str(cur_row)] = title

        cur_row += 1

    cur_row += 3
    sheet.row_dimensions[cur_row].height = 32
    sheet["F" + str(cur_row)].alignment = Alignment(wrap_text=True)
    sheet["F" + str(cur_row)] = "Total SKU Labor Cost"
    cur_row += 1
    sheet["F" + str(cur_row)] = round(total_sku_labeling_labor, 2)

    workbook.save(file_path)




def test_supply_levels(test_supplies_levels, supplies_needed_to_make_skus, item_quantity, full_sku):
    return_dict = {}
    return_dict["supplies_lacking"] = []

    quantity_able_to_make_list = []

    supplies_needed_for_one = supplies_needed_to_make_skus[full_sku]
    for supply_name in supplies_needed_for_one:
        quant_for_supply = supplies_needed_for_one[supply_name]
        tot = quant_for_supply * float(item_quantity)
        if supply_name not in test_supplies_levels:
            return_dict["supplies_lacking"].append(supply_name)
            quantity_able_to_make_list.append(0)
        else:
            test_new_supply_level = test_supplies_levels[supply_name] - tot
            if test_new_supply_level < 0:
                return_dict["supplies_lacking"].append(supply_name)
                if test_new_supply_level <= (-1 * item_quantity * quant_for_supply):
                    quantity_able_to_make_list.append(0)
                else:
                    quantity_able_to_make_list.append(item_quantity + (test_new_supply_level / quant_for_supply))

    quantity_able_to_make = item_quantity
    for val in quantity_able_to_make_list:
        if val < quantity_able_to_make:
            quantity_able_to_make = val
    return_dict["quantity_able_to_make"] = quantity_able_to_make

    if quantity_able_to_make > 0:
        for supply_name in supplies_needed_for_one:
            quant_for_supply = supplies_needed_for_one[supply_name]
            tot = quant_for_supply * float(quantity_able_to_make)
            test_supplies_levels[supply_name] = test_supplies_levels[supply_name] - tot

    return return_dict


def make_total_summary_excel(name_of_shipment_folder):
    root_path_to_info = loc_of_shipment_info + "/" + name_of_shipment_folder
    shipment_names = [
        f for f in os.listdir(root_path_to_info) if os.path.isdir(os.path.join(root_path_to_info, f))
    ]

    sku_to_info_for_excel_sheet = {}

    total_sku_labor_from_totals = 0

    for shipment in shipment_names:
        path_to_excel = root_path_to_info + "/" + shipment + "/summary_" + shipment + ".xlsx"
        workbook = openpyxl.load_workbook(path_to_excel)
        sheet = workbook.active
        cur_row = 2
        while True:
            sku = sheet["A" + str(cur_row)].value
            if sku == None:
                break

            carton_name = sheet["B" + str(cur_row)].value
            carton_quantity = int(sheet["C" + str(cur_row)].value)
            products_per_carton = int(sheet["D" + str(cur_row)].value)
            product_quantity = int(sheet["E" + str(cur_row)].value)
            sku_labor_cost_per_item = sheet["F" + str(cur_row)].value
            if sku_labor_cost_per_item != None:
                sku_labor_cost_per_item = float(sku_labor_cost_per_item)
            title = sheet["H" + str(cur_row)].value

            if sku in sku_to_info_for_excel_sheet:
                exist_carton_name = sku_to_info_for_excel_sheet[sku]["carton_name"]
                exist_products_per_carton = sku_to_info_for_excel_sheet[sku]["products_per_carton"]
                exist_sku_labor_cost_per_item = sku_to_info_for_excel_sheet[sku]["sku_labor_cost_per_item"]
                exist_title = sku_to_info_for_excel_sheet[sku]["title"]

                if exist_carton_name != carton_name:
                    raise Exception('exist_carton_name != carton_name for carton name: ' + str(carton_name))
                if exist_products_per_carton != products_per_carton:
                    raise Exception('exist_products_per_carton != products_per_carton for products_per_carton: ' + str(
                        products_per_carton))
                if exist_sku_labor_cost_per_item != sku_labor_cost_per_item:
                    raise Exception(
                        'exist_sku_labor_cost_per_item != sku_labor_cost_per_item for sku_labor_cost_per_item: ' + str(
                            sku_labor_cost_per_item))
                if exist_title != title:
                    raise Exception('exist_title != title for title: ' + str(title))

                exist_carton_quantity = sku_to_info_for_excel_sheet[sku]["carton_quantity"] + carton_quantity
                exist_product_quantity = sku_to_info_for_excel_sheet[sku]["product_quantity"] + product_quantity

                sku_to_info_for_excel_sheet[sku]["carton_quantity"] = exist_carton_quantity
                sku_to_info_for_excel_sheet[sku]["product_quantity"] = exist_product_quantity

            else:
                sku_to_info_for_excel_sheet[sku] = {}
                sku_to_info_for_excel_sheet[sku]["carton_name"] = carton_name
                sku_to_info_for_excel_sheet[sku]["carton_quantity"] = carton_quantity
                sku_to_info_for_excel_sheet[sku]["products_per_carton"] = products_per_carton
                sku_to_info_for_excel_sheet[sku]["product_quantity"] = product_quantity
                sku_to_info_for_excel_sheet[sku]["sku_labor_cost_per_item"] = sku_labor_cost_per_item
                sku_to_info_for_excel_sheet[sku]["title"] = title

            cur_row += 1
        row_of_SKU_labor_total = cur_row + 4
        total_sku_labor = sheet["F" + str(row_of_SKU_labor_total)].value
        if total_sku_labor != None:
            total_sku_labor = float(total_sku_labor)
        else:
            total_sku_labor = 0
        total_sku_labor_from_totals += total_sku_labor

    workbook = openpyxl.Workbook()
    view = [BookView(xWindow=0, yWindow=0, windowWidth=excel_window_width, windowHeight=excel_window_height)]
    workbook.views = view
    sheet = workbook.active
    sheet.sheet_view.zoomScale = 125
    sheet.protection.sheet = True

    sheet["A1"] = "SKU"
    sheet["B1"] = "Carton Name"
    sheet["C1"] = "Carton Quantity"
    sheet["D1"] = "Products Per Carton"
    sheet["E1"] = "Product Quantity"
    sheet["F1"] = "SKU Labor Cost Per Item"
    sheet["G1"] = "Location in Bays"
    sheet["H1"] = "Title"
    sheet['A1'].alignment = Alignment(wrap_text=True)
    sheet['B1'].alignment = Alignment(wrap_text=True)
    sheet['C1'].alignment = Alignment(wrap_text=True)
    sheet['D1'].alignment = Alignment(wrap_text=True)
    sheet['E1'].alignment = Alignment(wrap_text=True)
    sheet['F1'].alignment = Alignment(wrap_text=True)
    sheet['G1'].alignment = Alignment(wrap_text=True)
    sheet['H1'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[1].height = 32
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12

    total_sku_labeling_labor = 0

    cur_row = 2
    for sku in sku_to_info_for_excel_sheet:
        data = sku_to_info_for_excel_sheet[sku]
        carton_name = data["carton_name"]
        carton_quantity = data["carton_quantity"]
        products_per_carton = data["products_per_carton"]
        product_quantity = data["product_quantity"]
        sku_labor_cost_per_item = data["sku_labor_cost_per_item"]
        title = data["title"]

        location_in_bays = ""
        our_sku = carton_name.replace("C", "").strip()
        if our_sku in sku_to_location_in_bays:
            location_in_bays = sku_to_location_in_bays[our_sku]

        if (carton_quantity * products_per_carton) != product_quantity:
            raise Exception('(carton_quantity * products_per_carton) != product_quantity for sku: ' + str(sku))

        sheet["A" + str(cur_row)] = sku
        sheet["B" + str(cur_row)] = carton_name
        sheet["C" + str(cur_row)] = carton_quantity
        sheet["D" + str(cur_row)] = products_per_carton
        sheet["E" + str(cur_row)] = product_quantity
        if sku_labor_cost_per_item != None:
            sheet["F" + str(cur_row)] = sku_labor_cost_per_item
            total_sku_labeling_labor += (sku_labor_cost_per_item * product_quantity)
        sheet["G" + str(cur_row)] = location_in_bays
        sheet["H" + str(cur_row)] = title

        cur_row += 1

    cur_row += 3
    sheet.row_dimensions[cur_row].height = 32
    sheet["F" + str(cur_row)].alignment = Alignment(wrap_text=True)
    sheet["F" + str(cur_row)] = "Total SKU Labor Cost"
    cur_row += 1
    sheet["F" + str(cur_row)] = round(total_sku_labeling_labor, 2)

    percent_diff_between_labor_tots = abs(
        total_sku_labeling_labor - total_sku_labor_from_totals) / total_sku_labor_from_totals

    if percent_diff_between_labor_tots >= 0.01:
        raise Exception('percent_diff_between_labor_tots >= 0.01')

    workbook.save(root_path_to_info + "/total_summary.xlsx")


if __name__ == '__main__':
    get_aws_cred()
    get_access_token()

    before_run()

    make_excel_temp()

    # read_excel()

    after_run()
